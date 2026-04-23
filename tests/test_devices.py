from io import BytesIO

import openpyxl

from app import EXPORT_FIELDS, EXPORT_HEADERS, DEVICE_UPLOADS_DIR
from tests import BaseTestCase, db, json, os, shutil, _test_dir, patch


class TestDeviceCRUD(BaseTestCase):
    """Test device create, read, update, and lookup."""

    def test_add_and_get_device(self):
        device_id = db.add_device({'name': 'HP LaserJet', 'category': 'Printer'})
        device = db.get_device(device_id)
        self.assertIsNotNone(device)
        self.assertEqual(device['name'], 'HP LaserJet')
        self.assertEqual(device['category'], 'Printer')

    def test_add_device_with_hw_version(self):
        device_id = db.add_device({
            'name': 'HP Router',
            'category': 'Connectivity Device',
            'manufacturer': 'HP',
            'model_number': 'AX1000',
            'hw_version': 'Rev B',
        })
        device = db.get_device(device_id)
        self.assertEqual(device['hw_version'], 'Rev B')

    def test_update_device(self):
        device_id = db.add_device({'name': 'Old Name'})
        db.update_device(device_id, {'name': 'New Name'})
        device = db.get_device(device_id)
        self.assertEqual(device['name'], 'New Name')

    def test_get_device_by_barcode(self):
        device_id = db.add_device({'name': 'Scanner Test'})
        device = db.get_device(device_id)
        found = db.get_device_by_barcode(device['barcode_value'])
        self.assertIsNotNone(found)
        self.assertEqual(found['device_id'], device_id)

    def test_get_device_by_barcode_case_insensitive(self):
        device_id = db.add_device({'name': 'Case Test'})
        device = db.get_device(device_id)
        found = db.get_device_by_barcode(device['barcode_value'].lower())
        self.assertIsNotNone(found)

    def test_get_device_not_found(self):
        self.assertIsNone(db.get_device('nonexistent'))

    def test_get_device_by_barcode_not_found(self):
        self.assertIsNone(db.get_device_by_barcode('DOESNOTEXIST'))


class TestDuplicateSerialDetection(BaseTestCase):
    """Test duplicate serial number detection."""

    def test_get_device_by_serial(self):
        db.add_device({'name': 'Printer A', 'serial_number': 'SN12345'})
        found = db.get_device_by_serial('SN12345')
        self.assertIsNotNone(found)
        self.assertEqual(found['name'], 'Printer A')

    def test_get_device_by_serial_case_insensitive(self):
        db.add_device({'name': 'Printer B', 'serial_number': 'ABC123'})
        found = db.get_device_by_serial('abc123')
        self.assertIsNotNone(found)

    def test_retired_device_serial_not_found(self):
        device_id = db.add_device({'name': 'Printer C', 'serial_number': 'RET001'})
        db.retire_device(device_id)
        found = db.get_device_by_serial('RET001')
        self.assertIsNone(found)

    def test_duplicate_serial_blocked_in_ui(self):
        self.login_admin()
        # Add first device
        self.client.post('/devices/add', data={
            'manufacturer': 'HP', 'model_number': 'LJ100',
            'category': 'Router', 'serial_number': 'UNIQUE001',
        }, follow_redirects=True)
        # Try adding duplicate
        resp = self.client.post('/devices/add', data={
            'manufacturer': 'HP', 'model_number': 'LJ200',
            'category': 'Router', 'serial_number': 'UNIQUE001',
        }, follow_redirects=True)
        self.assertIn(b'already exists', resp.data)

    def test_custom_barcode_blocked_when_duplicate_in_ui(self):
        self.login_admin()
        self.client.post('/devices/add', data={
            'manufacturer': 'HP', 'model_number': 'LJ100',
            'category': 'Router', 'barcode_value': 'CNX-R900',
        }, follow_redirects=True)
        resp = self.client.post('/devices/add', data={
            'manufacturer': 'HP', 'model_number': 'LJ200',
            'category': 'Router', 'barcode_value': 'CNX-R900',
        }, follow_redirects=True)
        self.assertIn(b'barcode', resp.data.lower())
        self.assertIn(b'already exists', resp.data)


class TestDeviceEditRoute(BaseTestCase):
    """Test /devices/<id>/edit GET and POST."""

    def test_edit_form_loads(self):
        self.login_admin()
        did = db.add_device({'name': 'HP TestRouter', 'category': 'Connectivity Device', 'manufacturer': 'HP', 'model_number': 'TestRouter'})
        resp = self.client.get(f'/devices/{did}/edit')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'Edit Device', resp.data)

    def test_edit_updates_device(self):
        self.login_admin()
        did = db.add_device({'name': 'HP OldRouter', 'category': 'Connectivity Device', 'manufacturer': 'HP'})
        resp = self.client.post(f'/devices/{did}/edit', data={
            'manufacturer': 'Cisco', 'model_number': 'AX9000',
            'serial_number': 'SN-EDIT-001',
            'category': 'Connectivity Device', 'location': 'Lab B',
            'hw_version': 'Rev C',
        }, follow_redirects=True)
        self.assertIn(b'updated successfully', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['manufacturer'], 'Cisco')
        self.assertEqual(device['location'], 'Lab B')
        self.assertEqual(device['hw_version'], 'Rev C')

    def test_edit_updates_custom_barcode(self):
        self.login_admin()
        did = db.add_device({'name': 'HP OldRouter', 'category': 'Connectivity Device', 'manufacturer': 'HP'})
        resp = self.client.post(f'/devices/{did}/edit', data={
            'manufacturer': 'Cisco', 'model_number': 'AX9000',
            'barcode_value': 'CNX-R990',
            'category': 'Connectivity Device',
        }, follow_redirects=True)
        self.assertIn(b'updated successfully', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['barcode_value'], 'CNX-R990')

    def test_edit_nonexistent_device(self):
        self.login_admin()
        resp = self.client.get('/devices/nonexistent999/edit', follow_redirects=True)
        self.assertIn(b'Device not found', resp.data)

    def test_edit_requires_manufacturer_for_non_printer(self):
        self.login_admin()
        did = db.add_device({'name': 'HP Router', 'category': 'Connectivity Device', 'manufacturer': 'HP'})
        resp = self.client.post(f'/devices/{did}/edit', data={
            'manufacturer': '', 'category': 'Connectivity Device',
        }, follow_redirects=True)
        self.assertIn(b'Manufacturer is required', resp.data)

    def test_edit_printer_requires_codename(self):
        self.login_admin()
        did = db.add_device({'name': 'TestPrn (HP LJ)', 'category': 'Printer',
                             'manufacturer': 'HP', 'codename': 'TestPrn'})
        resp = self.client.post(f'/devices/{did}/edit', data={
            'manufacturer': 'HP', 'category': 'Printer', 'codename': '',
        }, follow_redirects=True)
        self.assertIn(b'Codename is required', resp.data)

    def test_edit_viewer_blocked(self):
        """Viewer cannot edit devices."""
        db.create_user('viewer1', 'pass1234', role='custom')
        did = db.add_device({'name': 'Locked Device'})
        self.client.post('/login', data={'username': 'viewer1', 'password': 'pass1234'})
        resp = self.client.post(f'/devices/{did}/edit', data={
            'manufacturer': 'HP', 'category': 'Connectivity Device',
        }, follow_redirects=True)
        self.assertIn(b'do not have permission', resp.data)


class TestDeviceBulkDelete(BaseTestCase):
    """Test /devices/bulk-delete."""

    def test_bulk_delete_removes_selected_devices(self):
        self.login_admin()
        d1 = db.add_device({'name': 'Delete Me 1', 'category': 'Connectivity Device'})
        d2 = db.add_device({'name': 'Delete Me 2', 'category': 'Connectivity Device'})
        keep = db.add_device({'name': 'Keep Me', 'category': 'Connectivity Device'})
        db.retire_device(d1)
        db.retire_device(d2)

        resp = self.client.post('/devices/bulk-delete', data={
            'device_ids': [d1, d2],
        }, follow_redirects=True)

        self.assertIn(b'Deleted 2 retired device(s).', resp.data)
        self.assertIsNone(db.get_device(d1))
        self.assertIsNone(db.get_device(d2))
        self.assertIsNotNone(db.get_device(keep))

    def test_bulk_delete_removes_retired_only_skips_active(self):
        self.login_admin()
        retired_id = db.add_device({'name': 'Old Device', 'category': 'Connectivity Device'})
        active_id = db.add_device({'name': 'Active Device', 'category': 'Connectivity Device'})
        db.retire_device(retired_id)

        resp = self.client.post('/devices/bulk-delete', data={
            'device_ids': [retired_id, active_id],
        }, follow_redirects=True)

        self.assertIn(b'Deleted 1 retired device(s).', resp.data)
        self.assertIn(b'1 selected device(s) were skipped', resp.data)
        self.assertIsNone(db.get_device(retired_id))
        self.assertIsNotNone(db.get_device(active_id))

    def test_bulk_delete_requires_devices_permission(self):
        db.create_user('viewer-bulk', 'pass1234', role='custom', permissions=[])
        did = db.add_device({'name': 'Protected Device', 'category': 'Connectivity Device'})
        self.client.post('/login', data={'username': 'viewer-bulk', 'password': 'pass1234'})

        resp = self.client.post('/devices/bulk-delete', data={'device_ids': did}, follow_redirects=True)

        self.assertIn(b'do not have permission', resp.data)
        self.assertIsNotNone(db.get_device(did))

    def test_bulk_delete_removes_device_upload_files(self):
        self.login_admin()
        did = db.add_device({'name': 'Has Attachment', 'category': 'Connectivity Device'})
        db.retire_device(did)
        upload_dir = os.path.join(DEVICE_UPLOADS_DIR, did)
        os.makedirs(upload_dir, exist_ok=True)
        filename = 'sample.txt'
        with open(os.path.join(upload_dir, filename), 'w', encoding='utf-8') as fh:
            fh.write('attachment')
        db.add_device_attachment(
            device_id=did,
            filename=filename,
            original_name='sample.txt',
            content_type='text/plain',
            size_bytes=10,
            uploaded_by='admin',
        )

        self.client.post('/devices/bulk-delete', data={'device_ids': did}, follow_redirects=True)

        self.assertIsNone(db.get_device(did))
        self.assertFalse(os.path.isdir(upload_dir))


class TestCheckoutCheckinFlow(BaseTestCase):
    """Test the full checkout/checkin lifecycle via web routes."""

    def test_checkout_device(self):
        self.login_admin()
        did = db.add_device({'name': 'Checkout Router'})
        resp = self.client.post(f'/devices/{did}/checkout', data={
            'assigned_to': 'John Doe',
        }, follow_redirects=True)
        self.assertIn(b'checked out to John Doe', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['status'], 'checked_out')
        self.assertEqual(device['assigned_to'], 'John Doe')

    def test_checkin_device(self):
        self.login_admin()
        did = db.add_device({'name': 'Checkin Router'})
        db.checkout_device(did, 'Jane Doe', performed_by='admin')
        resp = self.client.post(f'/devices/{did}/checkin', follow_redirects=True)
        self.assertIn(b'checked in', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['status'], 'available')
        self.assertEqual(device['assigned_to'], '')

    def test_checkout_empty_assignee_rejected(self):
        self.login_admin()
        did = db.add_device({'name': 'Empty Assign'})
        resp = self.client.post(f'/devices/{did}/checkout', data={
            'assigned_to': '',
        }, follow_redirects=True)
        self.assertIn(b'enter who', resp.data.lower())
        device = db.get_device(did)
        self.assertEqual(device['status'], 'available')

    def test_checkout_creates_audit_log(self):
        self.login_admin()
        did = db.add_device({'name': 'Audit Router'})
        self.client.post(f'/devices/{did}/checkout', data={
            'assigned_to': 'Auditor',
        }, follow_redirects=True)
        logs = db.get_audit_log(device_id=did)
        actions = [l['action'] for l in logs]
        self.assertIn('checked_out', actions)

    def test_checkin_creates_audit_log(self):
        self.login_admin()
        did = db.add_device({'name': 'Log Router'})
        db.checkout_device(did, 'Someone', performed_by='admin')
        self.client.post(f'/devices/{did}/checkin', follow_redirects=True)
        logs = db.get_audit_log(device_id=did)
        actions = [l['action'] for l in logs]
        self.assertIn('returned', actions)


class TestDeviceLifecycleFull(BaseTestCase):
    """Test complete device lifecycle: add -> edit -> checkout -> checkin -> retire."""

    def test_full_lifecycle(self):
        self.login_admin()
        # 1. Add
        resp = self.client.post('/devices/add', data={
            'manufacturer': 'HP', 'model_number': 'LaserJet 600',
            'serial_number': 'SN-LC-001',
            'category': 'Connectivity Device', 'location': 'Lab A',
        }, follow_redirects=True)
        self.assertIn(b'added successfully', resp.data)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        did = devices[0]['device_id']

        # 2. Edit
        resp = self.client.post(f'/devices/{did}/edit', data={
            'manufacturer': 'HP', 'model_number': 'LaserJet 601',
            'serial_number': 'SN-LC-001',
            'category': 'Connectivity Device', 'location': 'Lab B',
        }, follow_redirects=True)
        self.assertIn(b'updated successfully', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['location'], 'Lab B')

        # 3. Checkout
        resp = self.client.post(f'/devices/{did}/checkout', data={
            'assigned_to': 'Josh G',
        }, follow_redirects=True)
        self.assertIn(b'checked out', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['status'], 'checked_out')

        # 4. Checkin
        resp = self.client.post(f'/devices/{did}/checkin', follow_redirects=True)
        self.assertIn(b'checked in', resp.data)
        device = db.get_device(did)
        self.assertEqual(device['status'], 'available')

        # 5. Add note
        resp = self.client.post(f'/devices/{did}/notes', data={
            'note_content': 'Ready for retirement',
        }, follow_redirects=True)
        self.assertIn(b'Note added', resp.data)

        # 6. Retire
        resp = self.client.post(f'/devices/{did}/retire', data={
            'retire_reason': 'End of life cycle',
        }, follow_redirects=True)
        device = db.get_device(did)
        self.assertEqual(device['status'], 'retired')

        # 7. Verify full audit trail
        logs = db.get_audit_log(device_id=did)
        actions = [l['action'] for l in logs]
        self.assertIn('added', actions)
        self.assertIn('updated', actions)
        self.assertIn('checked_out', actions)
        self.assertIn('returned', actions)
        self.assertIn('retired', actions)

class TestDeviceNotes(BaseTestCase):
    """Test public device notes feature."""

    def _create_device(self):
        self.login_admin()
        did = db.add_device({'name': 'Note Test Device'})
        self.client.get('/logout')
        return did

    def test_notes_section_visible(self):
        """Device detail should show Notes section and add form."""
        did = self._create_device()
        resp = self.client.get(f'/devices/{did}')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'Notes', resp.data)
        self.assertIn(b'note_content', resp.data)

    def test_anonymous_add_note(self):
        """Anyone can add a note without logging in."""
        did = self._create_device()
        resp = self.client.post(f'/devices/{did}/notes', data={
            'note_content': 'Anonymous test note',
            'author_name': 'Tester Bob',
        }, follow_redirects=True)
        self.assertIn(b'Note added', resp.data)
        self.assertIn(b'Anonymous test note', resp.data)
        self.assertIn(b'Tester Bob', resp.data)

    def test_anonymous_default_name(self):
        """Omitting author_name defaults to 'Anonymous'."""
        did = self._create_device()
        self.client.post(f'/devices/{did}/notes', data={
            'note_content': 'No name note',
            'author_name': '',
        })
        notes = db.get_device_notes(did)
        self.assertEqual(notes[0]['author'], 'Anonymous')

    def test_logged_in_user_note(self):
        """Logged-in user's display name is used as author."""
        did = self._create_device()
        self.login_admin()
        resp = self.client.post(f'/devices/{did}/notes', data={
            'note_content': 'Admin note here',
        }, follow_redirects=True)
        self.assertIn(b'Note added', resp.data)
        self.assertIn(b'Admin note here', resp.data)

    def test_empty_note_rejected(self):
        """Empty notes should be rejected."""
        did = self._create_device()
        resp = self.client.post(f'/devices/{did}/notes', data={
            'note_content': '',
        }, follow_redirects=True)
        self.assertIn(b'cannot be empty', resp.data)

    def test_whitespace_only_note_rejected(self):
        """Whitespace-only notes should be rejected."""
        did = self._create_device()
        resp = self.client.post(f'/devices/{did}/notes', data={
            'note_content': '   \n  ',
        }, follow_redirects=True)
        self.assertIn(b'cannot be empty', resp.data)

    def test_too_long_note_rejected(self):
        """Notes over 2000 chars should be rejected."""
        did = self._create_device()
        resp = self.client.post(f'/devices/{did}/notes', data={
            'note_content': 'x' * 2001,
        }, follow_redirects=True)
        self.assertIn(b'too long', resp.data)

    def test_note_on_nonexistent_device(self):
        """Adding a note to a nonexistent device should fail gracefully."""
        resp = self.client.post('/devices/fake123/notes', data={
            'note_content': 'Orphan note',
        }, follow_redirects=True)
        self.assertIn(b'Device not found', resp.data)

    def test_admin_can_delete_note(self):
        """Admin can delete any note."""
        did = self._create_device()
        note_id = db.add_device_note(did, 'Bob', 'Delete me')
        self.login_admin()
        resp = self.client.post(f'/devices/{did}/notes/{note_id}/delete',
                                follow_redirects=True)
        self.assertIn(b'Note deleted', resp.data)
        self.assertEqual(len(db.get_device_notes(did)), 0)

    def test_non_admin_cannot_delete_note(self):
        """Non-admin users cannot delete notes."""
        did = self._create_device()
        note_id = db.add_device_note(did, 'Bob', 'Keep me')
        # Not logged in — should redirect
        resp = self.client.post(f'/devices/{did}/notes/{note_id}/delete',
                                follow_redirects=True)
        self.assertNotIn(b'Note deleted', resp.data)
        self.assertEqual(len(db.get_device_notes(did)), 1)

    def test_multiple_notes_ordered(self):
        """Multiple notes should all be returned."""
        did = self._create_device()
        db.add_device_note(did, 'Alice', 'First note')
        db.add_device_note(did, 'Bob', 'Second note')
        notes = db.get_device_notes(did)
        self.assertEqual(len(notes), 2)
        authors = {n['author'] for n in notes}
        self.assertIn('Alice', authors)
        self.assertIn('Bob', authors)


class TestDeviceNotesEdgeCases(BaseTestCase):
    """Test device notes edge cases."""

    def test_delete_nonexistent_note(self):
        self.login_admin()
        did = db.add_device({'name': 'Test Device'})
        resp = self.client.post(f'/devices/{did}/notes/99999/delete',
                                follow_redirects=True)
        self.assertEqual(resp.status_code, 200)

    def test_non_admin_cannot_delete_note(self):
        self.login_admin()
        did = db.add_device({'name': 'Note Delete Test'})
        db.add_device_note(did, 'Someone', 'A note')
        notes = db.get_device_notes(did)
        note_id = notes[0]['note_id']
        self.client.post('/users/add', data={
            'username': 'viewer1', 'password': 'test', 'role': 'custom',
            'permissions': ['wiki'],
        })
        self.client.get('/logout')
        self.client.post('/login', data={'username': 'viewer1', 'password': 'test'})
        resp = self.client.post(f'/devices/{did}/notes/{note_id}/delete',
                                follow_redirects=True)
        notes = db.get_device_notes(did)
        self.assertEqual(len(notes), 1)


class TestGetAllDevices(BaseTestCase):
    """Test get_all_devices include_retired flag."""

    def test_excludes_retired_by_default(self):
        db.add_device({'name': 'Active'})
        did2 = db.add_device({'name': 'Gone'})
        db.retire_device(did2)
        devices = db.get_all_devices()
        names = [d['name'] for d in devices]
        self.assertIn('Active', names)
        self.assertNotIn('Gone', names)

    def test_includes_retired_when_requested(self):
        db.add_device({'name': 'Active'})
        did2 = db.add_device({'name': 'Gone'})
        db.retire_device(did2)
        devices = db.get_all_devices(include_retired=True)
        names = [d['name'] for d in devices]
        self.assertIn('Active', names)
        self.assertIn('Gone', names)


class TestOwnershipDropdown(BaseTestCase):
    """Test the ownership dropdown (HP Owned / Vendor Supplied)."""

    def test_device_form_has_ownership_dropdown(self):
        self.login_admin()
        resp = self.client.get('/devices/add')
        self.assertIn(b'HP Owned', resp.data)
        self.assertIn(b'Vendor Supplied', resp.data)

    def test_vendor_supplied_persists(self):
        """Connectivity Devices can be set as vendor-supplied."""
        self.login_admin()
        self.client.post('/devices/add', data={
            'manufacturer': 'TP-Link', 'model_number': 'AX55',
            'serial_number': 'SN-001',
            'category': 'Connectivity Device', 'vendor_supplied': '1',
        }, follow_redirects=True)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        self.assertEqual(devices[0]['vendor_supplied'], 1)

    def test_hp_owned_default(self):
        """Non-Connectivity devices are always HP Owned regardless of form input."""
        self.login_admin()
        self.client.post('/devices/add', data={
            'manufacturer': 'HP', 'model_number': 'AX55',
            'serial_number': 'SN-002',
            'category': 'Endpoint Device', 'vendor_supplied': '1',
        }, follow_redirects=True)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        self.assertEqual(devices[0]['vendor_supplied'], 0)


class TestDeviceTypeAndMesh(BaseTestCase):
    """Test device_type dropdown and is_mesh checkbox."""

    def test_connectivity_device_type_persists(self):
        """device_type is stored for Connectivity Devices."""
        self.login_admin()
        self.client.post('/devices/add', data={
            'manufacturer': 'Cisco', 'model_number': 'MR46',
            'serial_number': 'SN-MR46-001',
            'category': 'Connectivity Device',
            'device_type': 'AP', 'is_mesh': '1',
        }, follow_redirects=True)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        self.assertEqual(devices[0]['device_type'], 'AP')
        self.assertEqual(devices[0]['is_mesh'], 1)

    def test_endpoint_device_type_persists(self):
        """device_type is stored for Endpoint Devices."""
        self.login_admin()
        self.client.post('/devices/add', data={
            'manufacturer': 'Dell', 'model_number': 'Latitude 5540',
            'serial_number': 'SN-DELL-001',
            'category': 'Endpoint Device',
            'device_type': 'Laptop',
        }, follow_redirects=True)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        self.assertEqual(devices[0]['device_type'], 'Laptop')
        self.assertEqual(devices[0]['is_mesh'], 0)

    def test_printer_device_type_defaults_na(self):
        """Printers get N/A for device_type."""
        self.login_admin()
        self.client.post('/devices/add', data={
            'category': 'Printer', 'codename': 'Cherry',
            'serial_number': 'SN-PR-001',
        }, follow_redirects=True)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        self.assertEqual(devices[0]['device_type'], 'N/A')


class TestDeviceValidation(BaseTestCase):
    """Test category-specific form validation."""

    def test_connectivity_requires_manufacturer(self):
        self.login_admin()
        resp = self.client.post('/devices/add', data={
            'category': 'Connectivity Device', 'model_number': 'AX55',
            'serial_number': 'SN-001',
        }, follow_redirects=True)
        self.assertIn(b'Manufacturer is required', resp.data)

    def test_connectivity_requires_model(self):
        self.login_admin()
        resp = self.client.post('/devices/add', data={
            'category': 'Connectivity Device', 'manufacturer': 'TP-Link',
            'serial_number': 'SN-001',
        }, follow_redirects=True)
        self.assertIn(b'Model number is required', resp.data)

    def test_connectivity_allows_empty_serial(self):
        self.login_admin()
        resp = self.client.post('/devices/add', data={
            'category': 'Connectivity Device', 'manufacturer': 'TP-Link',
            'model_number': 'AX55',
        }, follow_redirects=True)
        self.assertIn(b'added successfully', resp.data)

    def test_endpoint_requires_all_fields(self):
        self.login_admin()
        resp = self.client.post('/devices/add', data={
            'category': 'Endpoint Device',
        }, follow_redirects=True)
        self.assertIn(b'Manufacturer is required', resp.data)
        self.assertIn(b'Model number is required', resp.data)
        self.assertNotIn(b'Serial number is required', resp.data)

    def test_printer_defaults_hp_manufacturer(self):
        """Printer with empty manufacturer defaults to HP."""
        self.login_admin()
        self.client.post('/devices/add', data={
            'category': 'Printer', 'codename': 'Cherry',
            'serial_number': 'SN-HP-001',
        }, follow_redirects=True)
        devices = db.get_all_devices()
        self.assertEqual(len(devices), 1)
        self.assertEqual(devices[0]['manufacturer'], 'HP')


class TestDeviceViewedAudit(BaseTestCase):
    """Test that viewing a device creates an audit log entry."""

    def test_view_creates_audit_entry(self):
        device_id = db.add_device({'name': 'Audit Test'})
        self.client.get(f'/devices/{device_id}')
        log = db.get_audit_log(device_id=device_id)
        actions = [e['action'] for e in log]
        self.assertIn('viewed', actions)

    def test_scan_redirect_creates_scanned_action(self):
        """Device page loaded with ?scan=1 logs a 'scanned' action, not 'viewed'."""
        device_id = db.add_device({'name': 'Scan Test'})
        # The scan page appends ?scan=1 to the redirect URL
        resp = self.client.get(f'/devices/{device_id}?scan=1')
        self.assertEqual(resp.status_code, 200)
        log = db.get_audit_log(device_id=device_id)
        actions = [e['action'] for e in log]
        self.assertIn('scanned', actions)
        self.assertNotIn('viewed', actions)

    def test_plain_view_logs_viewed_not_scanned(self):
        """Device page loaded without ?scan=1 logs 'viewed'."""
        device_id = db.add_device({'name': 'View Test'})
        self.client.get(f'/devices/{device_id}')
        log = db.get_audit_log(device_id=device_id)
        actions = [e['action'] for e in log]
        self.assertIn('viewed', actions)
        self.assertNotIn('scanned', actions)


class TestDeviceAttachments(BaseTestCase):
    """Test device file attachment upload, download, preview, and delete."""

    def _make_device(self):
        self.login_admin()
        device_id = db.add_device({'name': 'Test Device', 'category': 'Printer'})
        return device_id

    def test_db_crud(self):
        """Database CRUD operations for device attachments."""
        device_id = db.add_device({'name': 'D1'})
        db.add_device_attachment(device_id, 'abc.png', 'photo.png', 'image/png', 1234, 'admin')
        atts = db.get_device_attachments(device_id)
        self.assertEqual(len(atts), 1)
        self.assertEqual(atts[0]['original_name'], 'photo.png')
        self.assertEqual(atts[0]['size_bytes'], 1234)

        att = db.get_device_attachment(atts[0]['attachment_id'])
        self.assertIsNotNone(att)
        self.assertEqual(att['filename'], 'abc.png')

        db.delete_device_attachment(atts[0]['attachment_id'])
        self.assertEqual(len(db.get_device_attachments(device_id)), 0)

    def test_upload_requires_login(self):
        """Upload should require authentication."""
        device_id = db.add_device({'name': 'D1'})
        resp = self.client.post(f'/devices/{device_id}/upload',
                                data={'attachment': (BytesIO(b'data'), 'test.txt')},
                                content_type='multipart/form-data',
                                follow_redirects=False)
        self.assertIn(resp.status_code, [302, 303])

    def test_upload_and_download(self):
        """Logged-in user can upload, anyone can download."""
        device_id = self._make_device()
        data = b'Hello attachment content'
        resp = self.client.post(f'/devices/{device_id}/upload',
                                data={'attachment': (BytesIO(data), 'readme.txt')},
                                content_type='multipart/form-data',
                                follow_redirects=True)
        self.assertEqual(resp.status_code, 200)

        atts = db.get_device_attachments(device_id)
        self.assertEqual(len(atts), 1)
        self.assertEqual(atts[0]['original_name'], 'readme.txt')

        # Download
        resp = self.client.get(f'/device/attachment/{atts[0]["attachment_id"]}')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data, data)

    def test_preview_image(self):
        """Image preview returns inline content."""
        device_id = self._make_device()
        png_data = b'\x89PNG\r\n\x1a\n' + b'\x00' * 100
        self.client.post(f'/devices/{device_id}/upload',
                         data={'attachment': (BytesIO(png_data), 'photo.png')},
                         content_type='multipart/form-data',
                         follow_redirects=True)
        atts = db.get_device_attachments(device_id)
        resp = self.client.get(f'/device/attachment/{atts[0]["attachment_id"]}/preview')
        self.assertEqual(resp.status_code, 200)

    def test_delete_requires_devices_permission(self):
        """Only users with devices permission can delete attachments."""
        device_id = self._make_device()
        self.client.post(f'/devices/{device_id}/upload',
                         data={'attachment': (BytesIO(b'x'), 'file.txt')},
                         content_type='multipart/form-data',
                         follow_redirects=True)
        atts = db.get_device_attachments(device_id)
        att_id = atts[0]['attachment_id']

        # Create a custom user without devices permission
        db.create_user('viewer', 'pass', role='custom', display_name='Viewer', permissions=[])
        self.client.get('/logout', follow_redirects=True)
        self.client.post('/login', data={'username': 'viewer', 'password': 'pass'}, follow_redirects=True)
        resp = self.client.post(f'/device/attachment/{att_id}/delete', follow_redirects=True)
        # Attachment should still exist
        self.assertIsNotNone(db.get_device_attachment(att_id))

        # Admin can delete
        self.client.get('/logout', follow_redirects=True)
        self.login_admin()
        resp = self.client.post(f'/device/attachment/{att_id}/delete', follow_redirects=True)
        self.assertEqual(resp.status_code, 200)
        self.assertIsNone(db.get_device_attachment(att_id))

    def test_rejected_extension(self):
        """Files with disallowed extensions are rejected."""
        device_id = self._make_device()
        resp = self.client.post(f'/devices/{device_id}/upload',
                                data={'attachment': (BytesIO(b'exe'), 'malware.exe')},
                                content_type='multipart/form-data',
                                follow_redirects=True)
        self.assertEqual(len(db.get_device_attachments(device_id)), 0)

    def test_attachments_shown_on_detail_page(self):
        """Device detail page includes attachments section."""
        device_id = self._make_device()
        self.client.post(f'/devices/{device_id}/upload',
                         data={'attachment': (BytesIO(b'content'), 'notes.txt')},
                         content_type='multipart/form-data',
                         follow_redirects=True)
        resp = self.client.get(f'/devices/{device_id}')
        self.assertIn(b'notes.txt', resp.data)
        self.assertIn(b'Attachments', resp.data)

    def test_download_nonexistent(self):
        """Downloading a nonexistent attachment returns 404."""
        resp = self.client.get('/device/attachment/99999')
        self.assertEqual(resp.status_code, 404)


def _xlsx_bytes_from_rows(data_rows):
    """Build .xlsx with export headers and optional data rows (values per EXPORT_FIELDS order)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([EXPORT_HEADERS[f] for f in EXPORT_FIELDS])
    for row in data_rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestDeviceImportXlsx(BaseTestCase):
    """POST /devices/import/xlsx — bulk import matching Export → Excel columns."""

    def test_import_redirects_to_login_when_anonymous(self):
        buf = _xlsx_bytes_from_rows([])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'inv.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_import_creates_connectivity_device(self):
        self.login_admin()
        row = [
            '', '', '',  # device_id, barcode, name (computed)
            'Connectivity Device',
            'N/A',
            'No',
            'Acme',
            'AP-900',
            'SN-XLSX-001',
            'Wi-Fi 6',
            'HP Owned',
            'available',
            'Shelf 1',
            '',
            '',
            'N/A',
            'N/A',
            '',
            '', '',
        ]
        self.assertEqual(len(row), len(EXPORT_FIELDS))
        buf = _xlsx_bytes_from_rows([row])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'batch.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)
        found = db.get_device_by_serial('SN-XLSX-001')
        self.assertIsNotNone(found)
        self.assertEqual(found['manufacturer'], 'Acme')
        self.assertEqual(found['model_number'], 'AP-900')
        self.assertEqual(found['location'], 'Shelf 1')

    def test_import_hyphen_with_nbsp_serial_normalized(self):
        """NBSP / odd whitespace around '-' still counts as empty serial."""
        self.login_admin()
        db.add_device({
            'name': 'Legacy Dash', 'category': 'Switch',
            'manufacturer': 'Old', 'model_number': 'M', 'serial_number': '-',
        })
        row = [
            '', '', '', 'Switch', 'N/A', 'No',
            'Acme', 'NBSP-SN', '\u00a0-\u00a0', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ]
        buf = _xlsx_bytes_from_rows([row])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'nbspsn.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)

    def test_import_hyphen_serial_not_treated_as_duplicate(self):
        """Spreadsheet '-' / N/A placeholders are empty serials, not a shared literal '-'."""
        self.login_admin()
        db.add_device({
            'name': 'Legacy Dash', 'category': 'Switch',
            'manufacturer': 'Old', 'model_number': 'M', 'serial_number': '-',
        })
        row = [
            '', '', '', 'Switch', 'N/A', 'No',
            'Acme', 'NEW-IMP', '—', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ]
        buf = _xlsx_bytes_from_rows([row])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'dashsn.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)
        matches = [
            d for d in db.get_all_devices()
            if d.get('model_number') == 'NEW-IMP' and d.get('manufacturer') == 'Acme'
        ]
        self.assertEqual(len(matches), 1)
        self.assertEqual((matches[0].get('serial_number') or '').strip(), '')

    def test_import_skips_duplicate_serial(self):
        self.login_admin()
        db.add_device({
            'name': 'Existing', 'category': 'Connectivity Device',
            'manufacturer': 'X', 'model_number': 'Y', 'serial_number': 'DUP-SN',
        })
        row = [
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'A', 'B', 'DUP-SN', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ]
        buf = _xlsx_bytes_from_rows([row])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'dup.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'already exists', resp.data)
        self.assertNotIn(b'Successfully imported 1', resp.data)

    def test_import_uses_custom_barcode_column_when_present(self):
        self.login_admin()
        row = [
            '', 'CNX-R777', '', 'Connectivity Device', 'N/A', 'No',
            'A', 'B', 'SN-CB-1', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ]
        buf = _xlsx_bytes_from_rows([row])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'cb.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)
        self.assertIsNotNone(db.get_device_by_barcode('CNX-R777'))

    def test_import_carries_down_empty_category(self):
        """Merged / empty Category cells inherit the previous row (like Excel fill-down)."""
        self.login_admin()
        row1 = [
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'Co', 'M1', 'SN-DOWN-01', '', 'HP Owned', 'available', 'L1', '', '', 'N/A', 'N/A', '', '',
        ]
        row2 = list(row1)
        row2[3] = ''
        row2[8] = 'SN-DOWN-02'
        buf = _xlsx_bytes_from_rows([row1, row2])
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'carry.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 2', resp.data)
        self.assertIsNotNone(db.get_device_by_serial('SN-DOWN-02'))

    def test_import_category_header_alias_zh(self):
        """Localized Category header (类别) maps to the category field."""
        self.login_admin()
        wb = openpyxl.Workbook()
        ws = wb.active
        hdr = [EXPORT_HEADERS[f] for f in EXPORT_FIELDS]
        hdr[3] = '类别'
        ws.append(hdr)
        ws.append([
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'Z', 'Z9', 'SN-ZH-HDR', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'zhhdr.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)
        self.assertIsNotNone(db.get_device_by_serial('SN-ZH-HDR'))

    def test_import_prefers_inventory_sheet_when_not_active(self):
        """Data on a sheet named Inventory is used even if another sheet is active."""
        self.login_admin()
        wb = openpyxl.Workbook()
        cover = wb.active
        cover.title = 'Cover'
        cover['A1'] = 'Title'
        inv = wb.create_sheet('Inventory')
        inv.append([EXPORT_HEADERS[f] for f in EXPORT_FIELDS])
        inv.append([
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'Z', 'Z9', 'SN-INV-SHEET', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ])
        wb.active = cover
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'sheets.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)
        self.assertIsNotNone(db.get_device_by_serial('SN-INV-SHEET'))

    def test_import_case_insensitive_headers(self):
        self.login_admin()
        wb = openpyxl.Workbook()
        ws = wb.active
        hdr = [EXPORT_HEADERS[f] for f in EXPORT_FIELDS]
        hdr[3] = 'category'
        ws.append(hdr)
        ws.append([
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'Z', 'Z9', 'SN-LOWHDR', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'lowhdr.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 1', resp.data)
        self.assertIsNotNone(db.get_device_by_serial('SN-LOWHDR'))

    def test_import_merged_category_column(self):
        """Merged Category cells only store value in top-left; import reads merge master."""
        self.login_admin()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([EXPORT_HEADERS[f] for f in EXPORT_FIELDS])
        base = [
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'M', 'M1', 'SN-MRG-1', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ]
        ws.append(base)
        r2 = list(base)
        r2[3] = None
        r2[8] = 'SN-MRG-2'
        ws.append(r2)
        r3 = list(base)
        r3[3] = None
        r3[8] = 'SN-MRG-3'
        ws.append(r3)
        ws.merge_cells(start_row=2, start_column=4, end_row=4, end_column=4)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (buf, 'merged.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 3', resp.data)
        self.assertIsNotNone(db.get_device_by_serial('SN-MRG-3'))

    def test_import_category_carries_over_blank_row(self):
        """Empty data rows no longer clear Category carry-down."""
        self.login_admin()
        row_a = [
            '', '', '', 'Connectivity Device', 'N/A', 'No',
            'A', 'A1', 'SN-BLANKGAP-1', '', 'HP Owned', 'available', '', '', '', 'N/A', 'N/A', '', '',
        ]
        row_b = list(row_a)
        row_b[3] = ''
        row_b[8] = 'SN-BLANKGAP-2'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([EXPORT_HEADERS[f] for f in EXPORT_FIELDS])
        ws.append(row_a)
        ws.append([None] * len(EXPORT_FIELDS))
        ws.append(row_b)
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = self.client.post(
            '/devices/import/xlsx',
            data={'file': (out, 'gap.xlsx')},
            content_type='multipart/form-data',
            follow_redirects=True,
        )
        self.assertIn(b'Successfully imported 2', resp.data)
        self.assertIsNotNone(db.get_device_by_serial('SN-BLANKGAP-2'))
