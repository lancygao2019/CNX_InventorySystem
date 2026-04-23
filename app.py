"""
HP Connectivity Team Inventory Management System — Flask Application

All routes are defined here. Run with: python app.py [--host HOST] [--port PORT]

Authentication: Admins (scanner terminal) can add/edit/checkout/retire/import devices.
Anyone on the network can view the inventory without logging in.
"""

import argparse
import csv
import io
import json
import logging
import os
import re
import subprocess
import sys
import traceback
import uuid
import zipfile
from functools import wraps
from datetime import datetime, timedelta, timezone
from logging.handlers import RotatingFileHandler

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, Response, session, g,
)
from runtime_dirs import BUNDLE_DIR, DATA_DIR, GIT_EXECUTABLE

import database as db
import barcode_utils

app = Flask(__name__,
            static_folder=os.path.join(BUNDLE_DIR, 'static'),
            template_folder=os.path.join(BUNDLE_DIR, 'templates'))
app.secret_key = os.environ.get('SECRET_KEY', 'hp-connectivity-inventory-system-change-me')
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 3600  # cache static files for 1 hour

# Application version (read from VERSION file)
# In PyInstaller bundles, look in BUNDLE_DIR (sys._MEIPASS); otherwise look
# next to app.py. Also check DATA_DIR as a fallback for dev-mode runs.
_version_candidates = [
    os.path.join(BUNDLE_DIR, 'VERSION'),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'VERSION'),
    os.path.join(DATA_DIR, 'VERSION'),
]
_app_version = 'dev'
for _vp in _version_candidates:
    try:
        with open(_vp) as _vf:
            _app_version = _vf.read().strip()
            break
    except (FileNotFoundError, OSError):
        continue

# ---------------------------------------------------------------------------
# Application logging (rotating file, single file that overwrites at limit)
# ---------------------------------------------------------------------------

LOG_DIR = os.path.join(DATA_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, 'app.log')
LOG_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'log_config.json')
SERVER_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'server_config.json')


def _load_log_config():
    try:
        with open(LOG_CONFIG_FILE, 'r') as f:
            import json as _j
            return _j.load(f)
    except (FileNotFoundError, ValueError):
        return {'max_size_mb': 2}


def _save_log_config(config):
    import json as _j
    with open(LOG_CONFIG_FILE, 'w') as f:
        _j.dump(config, f)


_DEFAULT_UPDATE_REPO_URL = 'https://github.com/jgaweda/Inventory-Management'


def _load_server_config():
    # Default values. update_repo_url is stored as empty string when using
    # the built-in default — this keeps the admin UI field blank so users
    # only see their own custom value (if any).
    defaults = {
        'port': 8080,
        'host': '0.0.0.0',
        'update_repo_url': '',
        'update_branch': '',  # empty = use releases instead of a branch
    }
    try:
        with open(SERVER_CONFIG_FILE, 'r') as f:
            saved = json.load(f)
    except (FileNotFoundError, ValueError):
        saved = {}
    defaults.update(saved)
    return defaults


def _get_update_repo_url(config=None):
    """Resolve the effective update repo URL: user override or built-in default."""
    if config is None:
        config = _load_server_config()
    return (config.get('update_repo_url') or '').strip() or _DEFAULT_UPDATE_REPO_URL


def _save_server_config(config):
    with open(SERVER_CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)


_AUTOSTART_REGISTRY_KEY = r'Software\Microsoft\Windows\CurrentVersion\Run'
_AUTOSTART_VALUE_NAME = 'HPConnectivityInventory'


def _get_autostart_enabled():
    """Return True if the Windows autostart registry entry is set."""
    if sys.platform != 'win32':
        return False
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _AUTOSTART_REGISTRY_KEY) as key:
            winreg.QueryValueEx(key, _AUTOSTART_VALUE_NAME)
            return True
    except (ImportError, FileNotFoundError, OSError):
        return False


def _set_autostart(enabled):
    """Enable or disable running on startup via Windows registry (HKCU\\...\\Run).
    Returns (ok, message)."""
    if sys.platform != 'win32':
        return False, 'Autostart is only supported on Windows.'
    try:
        import winreg
    except ImportError:
        return False, 'winreg module is not available on this platform.'

    # Determine the executable path: frozen exe or python interpreter + script
    if getattr(sys, 'frozen', False):
        exe_path = sys.executable
    else:
        exe_path = f'"{sys.executable}" "{os.path.abspath(__file__)}"'

    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _AUTOSTART_REGISTRY_KEY,
                            0, winreg.KEY_SET_VALUE) as key:
            if enabled:
                # Frozen exe is windowed (no console) and runs silently.
                # For source runs, wrap in cmd /c start /min so there's no
                # lingering terminal window.
                if getattr(sys, 'frozen', False):
                    command = f'"{exe_path}"'
                else:
                    command = f'cmd /c start "" /min {exe_path}'
                winreg.SetValueEx(key, _AUTOSTART_VALUE_NAME, 0, winreg.REG_SZ, command)
                return True, 'Autostart enabled — the application will launch on user login.'
            else:
                try:
                    winreg.DeleteValue(key, _AUTOSTART_VALUE_NAME)
                except FileNotFoundError:
                    pass
                return True, 'Autostart disabled.'
    except OSError as e:
        return False, f'Failed to update registry: {e}'


_log_config = _load_log_config()
_log_max_bytes = int(_log_config.get('max_size_mb', 2) * 1024 * 1024)

app_logger = logging.getLogger('inventory')
app_logger.setLevel(logging.DEBUG)
_log_handler = RotatingFileHandler(LOG_FILE, maxBytes=_log_max_bytes, backupCount=1)
_log_handler.setFormatter(logging.Formatter(
    '%(asctime)s | %(levelname)-7s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S'
))
app_logger.addHandler(_log_handler)


def _reconfigure_log_handler(max_size_mb):
    """Update the log handler's max size at runtime."""
    _log_handler.maxBytes = int(max_size_mb * 1024 * 1024)

# ---------------------------------------------------------------------------
# Startup: initialize the database
# ---------------------------------------------------------------------------

with app.app_context():
    db.init_db()
    os.makedirs(os.path.join(app.static_folder, 'labels'), exist_ok=True)
    os.makedirs(db._get_backup_dir(), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, 'wiki_uploads'), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, 'device_uploads'), exist_ok=True)
    app_logger.info('Application started — database initialized')

    # Defer slow integrity checks to a background thread so the server
    # starts accepting requests immediately (big win on slow machines).
    import threading

    def _deferred_startup_checks():
        try:
            _integrity = db.startup_integrity_check()
            if not _integrity['ok']:
                app_logger.error('DATABASE INTEGRITY ISSUE: %s', _integrity['result'])
            _att_check = db.check_attachment_integrity(os.path.join(DATA_DIR, 'wiki_uploads'))
            if _att_check['orphaned_removed'] > 0:
                app_logger.warning('Removed %d orphaned wiki attachment records', _att_check['orphaned_removed'])
        except Exception as e:
            app_logger.error('Deferred startup check failed: %s', e)

    threading.Thread(target=_deferred_startup_checks, daemon=True).start()

# ---------------------------------------------------------------------------
# Authentication helpers
# ---------------------------------------------------------------------------

@app.before_request
def load_user():
    """Load the current user from session before each request."""
    g.user = None
    user_id = session.get('user_id')
    if user_id:
        g.user = db.get_user(user_id)
        if not g.user:
            session.clear()


# Periodically check if the scheduler thread is alive (every ~60 seconds)
_last_scheduler_check = datetime.now()


@app.before_request
def _check_scheduler_health():
    """Self-heal: restart scheduler thread if it died, checked at most once per minute."""
    global _last_scheduler_check
    now = datetime.now()
    if (now - _last_scheduler_check).total_seconds() < 60:
        return
    _last_scheduler_check = now
    if _scheduler_thread is not None and not _scheduler_thread.is_alive():
        app_logger.warning('Scheduler thread found dead — restarting')
        _ensure_scheduler_running()


def login_required(f):
    """Decorator: redirect to login if not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not g.user:
            flash('Please log in to continue.', 'warning')
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated


# Centralized permission model — single source of truth for all role access.
# To change what a role can do, edit this dict. To add a role, add a line.
ROLE_PERMISSIONS = {
    'admin':  {'devices', 'references', 'wiki', 'users', 'backups', 'logs', 'settings', 'retire', 'update'},
    'custom': set(),  # custom users get permissions from their user record
}

# Assignable permissions shown as checkboxes when creating/editing custom users.
# Admin-only permissions (users, settings) are not assignable.
ASSIGNABLE_PERMISSIONS = [
    ('devices',      'Devices — Add, edit, checkout/checkin, and delete notes'),
    ('references',   'References — Manage product reference catalog'),
    ('wiki',         'Wiki — Edit pages, upload/delete attachments'),
    ('backups',      'Backups — View, create, restore, and configure backups'),
    ('logs',         'Logs — View and export application logs'),
    ('retire',       'Retire — Retire and unretire devices'),
    ('update',       'Update — Check for and apply application updates'),
]

# Permissions that can be granted to non-logged-in (guest/public) users.
# Sensitive permissions (backups, logs, settings, users) are excluded.
GUEST_ASSIGNABLE_PERMISSIONS = [
    ('references',   'References — Manage product reference catalog'),
    ('wiki',         'Wiki — Edit pages, upload/delete attachments'),
]


def get_user_permissions(user):
    """Return the effective permission set for a user dict.
    For guest (None) users, returns permissions from the database config."""
    if not user:
        return db.get_guest_permissions()
    if user['role'] == 'admin':
        return ROLE_PERMISSIONS['admin']
    # Custom users: permissions is a pre-parsed list from _parse_user_row
    perms = user.get('permissions')
    if isinstance(perms, list):
        result = set(perms)
    elif isinstance(perms, str):
        try:
            result = set(json.loads(perms))
        except (json.JSONDecodeError, TypeError):
            result = set()
    else:
        result = set()
    # Migrate legacy permissions: wiki_admin → wiki, notes_delete → devices
    if 'wiki_admin' in result:
        result.discard('wiki_admin')
        result.add('wiki')
    if 'notes_delete' in result:
        result.discard('notes_delete')
        result.add('devices')
    return result


def has_permission(permission):
    """Check if the current user (or guest) has a specific permission."""
    return permission in get_user_permissions(g.user)


def permission_required(permission):
    """Decorator: require a specific permission. Guests with the permission are allowed through."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if has_permission(permission):
                return f(*args, **kwargs)
            if not g.user:
                flash('Please log in to continue.', 'warning')
                return redirect(url_for('login', next=request.path))
            flash('You do not have permission to perform this action.', 'error')
            return redirect(url_for('dashboard'))
        return decorated
    return decorator


def current_username():
    """Return display name of logged-in user, or 'guest'."""
    if g.user:
        return g.user['display_name'] or g.user['username']
    return 'guest'

# ---------------------------------------------------------------------------
# Context processor: inject categories, user, and current time into templates
# ---------------------------------------------------------------------------

@app.context_processor
def inject_globals():
    return {
        'categories': db.get_categories(),
        'now': datetime.now(timezone.utc),
        'current_user': g.user,
        'has_permission': has_permission,
        'app_version': _app_version,
        'backup_health': db.get_backup_health(),
    }

# ---------------------------------------------------------------------------
# Login / Logout
# ---------------------------------------------------------------------------

# Simple in-memory rate limiter for login
_login_attempts = {}  # ip -> [timestamp, ...]
_LOGIN_WINDOW = 300   # 5 minutes
_LOGIN_MAX = 10       # max attempts per window


@app.route('/login', methods=['GET', 'POST'])
def login():
    if g.user:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        ip = request.remote_addr
        now = datetime.now().timestamp()
        # Clean old attempts and check rate
        attempts = [t for t in _login_attempts.get(ip, []) if now - t < _LOGIN_WINDOW]
        if len(attempts) >= _LOGIN_MAX:
            app_logger.warning('Login rate limited: ip=%s attempts=%d', ip, len(attempts))
            flash('Too many login attempts. Please wait a few minutes.', 'error')
            return render_template('login.html', next=request.args.get('next', ''))

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = db.authenticate_user(username, password)
        if user:
            _login_attempts.pop(ip, None)  # Clear on success
            session['user_id'] = user['user_id']
            session['role'] = user['role']
            app_logger.info('Login successful: user=%s role=%s ip=%s', username, user['role'], request.remote_addr)
            next_url = request.form.get('next', '')
            if not next_url or next_url.startswith('//') or '://' in next_url:
                next_url = url_for('dashboard')
            return redirect(next_url)
        else:
            attempts.append(now)
            _login_attempts[ip] = attempts
            app_logger.warning('Login failed: user=%s ip=%s attempt=%d/%d', username, ip, len(attempts), _LOGIN_MAX)
            flash('Invalid username or password.', 'error')
            # Show password hint if one is set for this username
            hint = db.get_password_hint(username)
            if hint:
                flash(f'Hint: {hint}', 'warning')
            return render_template('login.html', next=request.form.get('next', ''))

    return render_template('login.html', next=request.args.get('next', ''))


@app.route('/logout')
def logout():
    username = current_username()
    session.clear()
    app_logger.info('Logout: user=%s ip=%s', username, request.remote_addr)
    flash('You have been logged out.', 'success')
    return redirect(url_for('dashboard'))

# ---------------------------------------------------------------------------
# Dashboard (public)
# ---------------------------------------------------------------------------

@app.route('/health')
def health():
    """Health check endpoint for monitoring and CI smoke tests."""
    try:
        integrity = db.check_database_integrity()
        return jsonify({'status': 'ok', 'db': integrity['result']})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/')
def dashboard():
    return redirect(url_for('scan_page'))

# ---------------------------------------------------------------------------
# Device list (public)
# ---------------------------------------------------------------------------

@app.route('/devices')
def device_list():
    q = request.args.get('q', '')
    selected_category = request.args.get('category', '')
    selected_status = request.args.get('status', '')
    selected_connectivity = request.args.get('connectivity', '')
    selected_location = request.args.get('location', '')
    codename = request.args.get('codename', '')
    try:
        per_page = int(request.args.get('per_page', 50))
    except (TypeError, ValueError):
        per_page = 50
    if per_page not in (50, 100):
        per_page = 50
    try:
        page = int(request.args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    page = max(page, 1)

    devices, total_devices = db.search_devices_paginated(
        query=q,
        category=selected_category,
        status=selected_status,
        connectivity=selected_connectivity,
        location=selected_location,
        codename=codename,
        limit=per_page,
        offset=(page - 1) * per_page
    )
    total_pages = max((total_devices + per_page - 1) // per_page, 1)
    if page > total_pages:
        page = total_pages
        devices, total_devices = db.search_devices_paginated(
            query=q,
            category=selected_category,
            status=selected_status,
            connectivity=selected_connectivity,
            location=selected_location,
            codename=codename,
            limit=per_page,
            offset=(page - 1) * per_page
        )
    start_index = ((page - 1) * per_page + 1) if total_devices else 0
    end_index = min(page * per_page, total_devices) if total_devices else 0

    return render_template(
        'devices.html',
        devices=devices,
        q=q,
        selected_category=selected_category,
        selected_status=selected_status,
        selected_connectivity=selected_connectivity,
        selected_location=selected_location,
        selected_codename=codename,
        page=page,
        per_page=per_page,
        per_page_options=(50, 100),
        total_pages=total_pages,
        total_devices=total_devices,
        start_index=start_index,
        end_index=end_index
    )

# ---------------------------------------------------------------------------
# Add device (admin only)
# ---------------------------------------------------------------------------

@app.route('/devices/add', methods=['GET', 'POST'])
@permission_required('devices')
def device_add():
    if request.method == 'POST':
        manufacturer = request.form.get('manufacturer', '').strip()
        model_number = request.form.get('model_number', '').strip()
        category = request.form.get('category', '').strip()
        codename = request.form.get('codename', '').strip()

        # "Other" with custom detail becomes "Other - <detail>"
        if category == 'Other':
            other_detail = request.form.get('other_detail', '').strip()
            if other_detail:
                category = f'Other - {other_detail}'

        serial_number = request.form.get('serial_number', '').strip()
        custom_barcode = _normalize_custom_barcode(request.form.get('barcode_value', ''))
        custom_barcode = _normalize_custom_barcode(request.form.get('barcode_value', ''))
        custom_barcode = _normalize_custom_barcode(request.form.get('barcode_value', ''))

        # Category-specific validation
        errors = []
        if category == 'Printer':
            if not manufacturer:
                manufacturer = 'HP'
            if not codename:
                errors.append('Codename is required for printers.')
        elif category in ('Connectivity Device', 'Endpoint Device'):
            if not manufacturer:
                errors.append('Manufacturer is required.')
            if not model_number:
                errors.append('Model number is required.')

        if errors:
            for e in errors:
                flash(e, 'error')
            return render_template('device_form.html', device=request.form, is_edit=False)
        bc_err = _validate_custom_barcode_or_error(custom_barcode)
        if bc_err:
            flash(bc_err, 'error')
            return render_template('device_form.html', device=request.form, is_edit=False)

        if category == 'Printer':
            variant = request.form.get('variant', '').strip()
            codename_display = f'{codename} {variant}'.strip() if variant else codename
            mfg_model = f'{manufacturer} {model_number}'.strip()
            name = f'{codename_display} ({mfg_model})' if mfg_model else codename_display
        else:
            name = f'{manufacturer} {model_number}'.strip()
        if serial_number:
            existing = db.get_device_by_serial(serial_number)
            if existing:
                flash(f'A device with serial number "{serial_number}" already exists: {existing["name"]}', 'error')
                return render_template('device_form.html', device=request.form, is_edit=False)
        if custom_barcode:
            existing_bc = db.get_device_by_barcode(custom_barcode)
            if existing_bc:
                flash(f'A device with barcode "{custom_barcode}" already exists: {existing_bc["name"]}', 'error')
                return render_template('device_form.html', device=request.form, is_edit=False)

        # Only Connectivity Devices can be vendor-supplied; all others are HP Owned
        vendor_supplied = 1 if (category == 'Connectivity Device' and request.form.get('vendor_supplied') == '1') else 0

        data = {
            'name': name,
            'category': category,
            'barcode_value': custom_barcode,
            'manufacturer': manufacturer,
            'model_number': model_number,
            'serial_number': serial_number,
            'connectivity': request.form.get('connectivity', ''),
            'hw_version': request.form.get('hw_version', '').strip(),
            'vendor_supplied': vendor_supplied,
            'location': request.form.get('location', ''),
            'notes': request.form.get('notes', ''),
            'codename': codename or ('N/A' if category != 'Printer' else ''),
            'variant': request.form.get('variant', '').strip() or ('N/A' if category != 'Printer' else ''),
            'device_type': request.form.get('device_type', '').strip() or ('N/A' if category not in ('Connectivity Device', 'Endpoint Device') else ''),
            'is_mesh': 1 if request.form.get('is_mesh') == '1' else 0,
        }
        device_id = db.add_device(data, performed_by=current_username())

        # Generate label
        device = db.get_device(device_id)
        barcode_utils.generate_label(device_id, device['barcode_value'], _label_name(device))

        app_logger.info('Device added: id=%s name="%s" cat="%s" by=%s', device_id, name, category, current_username())
        flash(f'Device "{name}" added successfully.', 'success')
        return redirect(url_for('device_detail', device_id=device_id))

    return render_template('device_form.html', device={}, is_edit=False)

# ---------------------------------------------------------------------------
# Device detail (public)
# ---------------------------------------------------------------------------

@app.route('/devices/<device_id>')
def device_detail(device_id):
    device = db.get_device(device_id)
    if not device:
        app_logger.warning('Device not found: id=%s ip=%s', device_id, request.remote_addr)
        flash('Device not found.', 'error')
        return redirect(url_for('device_list'))

    if not barcode_utils.label_exists(device_id):
        barcode_utils.generate_label(device_id, device['barcode_value'], _label_name(device))
        app_logger.debug('Label generated on-the-fly: id=%s', device_id)

    # Log the view/scan in the audit trail so it shows in Recent Activity.
    # The scan page appends ?scan=1 to the redirect URL so we can tell the
    # difference between a barcode scan and a regular page view.
    from_scan = request.args.get('scan') == '1'
    action = 'scanned' if from_scan else 'viewed'
    with db.db_transaction() as conn:
        db.log_action(conn, device_id, action, performed_by=current_username())
    audit = db.get_audit_log(device_id=device_id, limit=50)

    # Look up product reference data if this is a printer with a codename
    prod_ref = None
    if device.get('category') == 'Printer':
        codename = device.get('codename', '')
        if codename:
            refs = db.get_product_reference_by_codename(codename)
            if refs:
                prod_ref = refs[0]

    device_notes = db.get_device_notes(device_id)
    attachments = db.get_device_attachments(device_id)
    return render_template('device_detail.html', device=device, audit=audit,
                           prod_ref=prod_ref, device_notes=device_notes,
                           attachments=attachments)

# ---------------------------------------------------------------------------
# Device notes (public — anyone can add)
# ---------------------------------------------------------------------------

@app.route('/devices/<device_id>/notes', methods=['POST'])
def add_device_note(device_id):
    """Add a note to a device. Anyone can add notes."""
    device = db.get_device(device_id)
    if not device:
        flash('Device not found.', 'error')
        return redirect(url_for('device_list'))

    content = request.form.get('note_content', '').strip()
    if not content:
        flash('Note cannot be empty.', 'error')
        return redirect(url_for('device_detail', device_id=device_id))

    if len(content) > 2000:
        flash('Note is too long (max 2000 characters).', 'error')
        return redirect(url_for('device_detail', device_id=device_id))

    author = current_username() if g.user else request.form.get('author_name', '').strip()
    if not author:
        author = 'Anonymous'

    db.add_device_note(device_id, author, content)
    app_logger.info('Note added to device %s by %s', device_id, author)
    flash('Note added.', 'success')
    return redirect(url_for('device_detail', device_id=device_id))


@app.route('/devices/<device_id>/notes/<int:note_id>/delete', methods=['POST'])
@permission_required('devices')
def delete_device_note_route(device_id, note_id):
    """Delete a device note (requires devices permission)."""
    db.delete_device_note(note_id)
    app_logger.info('Note deleted: note_id=%d device_id=%s by=%s', note_id, device_id, current_username())
    flash('Note deleted.', 'success')
    return redirect(url_for('device_detail', device_id=device_id))

# ---------------------------------------------------------------------------
# Edit device (admin only)
# ---------------------------------------------------------------------------

@app.route('/devices/<device_id>/edit', methods=['GET', 'POST'])
@permission_required('devices')
def device_edit(device_id):
    device = db.get_device(device_id)
    if not device:
        flash('Device not found.', 'error')
        return redirect(url_for('device_list'))

    if request.method == 'POST':
        manufacturer = request.form.get('manufacturer', '').strip()
        model_number = request.form.get('model_number', '').strip()
        category = request.form.get('category', '').strip()
        codename = request.form.get('codename', '').strip()

        # "Other" with custom detail becomes "Other - <detail>"
        if category == 'Other':
            other_detail = request.form.get('other_detail', '').strip()
            if other_detail:
                category = f'Other - {other_detail}'

        serial_number = request.form.get('serial_number', '').strip()
        custom_barcode = _normalize_custom_barcode(request.form.get('barcode_value', ''))

        # Category-specific validation
        errors = []
        if category == 'Printer':
            if not manufacturer:
                manufacturer = 'HP'
            if not codename:
                errors.append('Codename is required for printers.')
        elif category in ('Connectivity Device', 'Endpoint Device'):
            if not manufacturer:
                errors.append('Manufacturer is required.')
            if not model_number:
                errors.append('Model number is required.')

        if errors:
            for e in errors:
                flash(e, 'error')
            return render_template('device_form.html', device=request.form, is_edit=True, device_id=device_id)
        bc_err = _validate_custom_barcode_or_error(custom_barcode)
        if bc_err:
            flash(bc_err, 'error')
            return render_template('device_form.html', device=request.form, is_edit=True, device_id=device_id)

        if category == 'Printer':
            variant = request.form.get('variant', '').strip()
            codename_display = f'{codename} {variant}'.strip() if variant else codename
            mfg_model = f'{manufacturer} {model_number}'.strip()
            name = f'{codename_display} ({mfg_model})' if mfg_model else codename_display
        else:
            name = f'{manufacturer} {model_number}'.strip()
        if serial_number and serial_number != (device.get('serial_number') or '').strip():
            existing = db.get_device_by_serial(serial_number)
            if existing and existing.get('device_id') != device_id:
                flash(f'A device with serial number "{serial_number}" already exists: {existing["name"]}', 'error')
                return render_template('device_form.html', device=request.form, is_edit=True, device_id=device_id)
        current_bc = (device.get('barcode_value') or '').strip().upper()
        if custom_barcode and custom_barcode != current_bc:
            existing_bc = db.get_device_by_barcode(custom_barcode)
            if existing_bc and existing_bc.get('device_id') != device_id:
                flash(f'A device with barcode "{custom_barcode}" already exists: {existing_bc["name"]}', 'error')
                return render_template('device_form.html', device=request.form, is_edit=True, device_id=device_id)

        # Only Connectivity Devices can be vendor-supplied; all others are HP Owned
        vendor_supplied = 1 if (category == 'Connectivity Device' and request.form.get('vendor_supplied') == '1') else 0

        data = {
            'name': name,
            'category': category,
            'barcode_value': custom_barcode or current_bc,
            'manufacturer': manufacturer,
            'model_number': model_number,
            'serial_number': serial_number,
            'connectivity': request.form.get('connectivity', ''),
            'hw_version': request.form.get('hw_version', '').strip(),
            'vendor_supplied': vendor_supplied,
            'status': request.form.get('status', device['status']),
            'location': request.form.get('location', ''),
            'assigned_to': request.form.get('assigned_to', ''),
            'notes': request.form.get('notes', ''),
            'codename': codename or ('N/A' if category != 'Printer' else ''),
            'variant': request.form.get('variant', '').strip() or ('N/A' if category != 'Printer' else ''),
            'device_type': request.form.get('device_type', '').strip() or ('N/A' if category not in ('Connectivity Device', 'Endpoint Device') else ''),
            'is_mesh': 1 if request.form.get('is_mesh') == '1' else 0,
        }
        db.update_device(device_id, data, performed_by=current_username())

        updated_device = db.get_device(device_id)
        barcode_utils.generate_label(device_id, device['barcode_value'], _label_name(updated_device))

        app_logger.info('Device updated: id=%s name="%s" cat="%s" by=%s', device_id, name, category, current_username())
        flash(f'Device "{name}" updated successfully.', 'success')
        return redirect(url_for('device_detail', device_id=device_id))

    return render_template('device_form.html', device=device, is_edit=True, device_id=device_id)

# ---------------------------------------------------------------------------
# Retire device (admin only)
# ---------------------------------------------------------------------------

@app.route('/devices/<device_id>/retire', methods=['POST'])
@permission_required('retire')
def device_retire(device_id):
    reason = request.form.get('retire_reason', '').strip()
    if not reason:
        flash('A reason is required to retire a device.', 'error')
        return redirect(url_for('device_detail', device_id=device_id))
    db.retire_device(device_id, performed_by=current_username(), reason=reason)
    app_logger.info('Device retired: id=%s reason="%s" by=%s', device_id, reason, current_username())
    flash('Device retired successfully.', 'success')
    return redirect(url_for('device_list'))


@app.route('/devices/bulk-delete', methods=['POST'])
@permission_required('devices')
def device_bulk_delete():
    """Delete multiple retired devices selected from the inventory list."""
    raw_ids = []
    for item in request.form.getlist('device_ids'):
        raw_ids.extend(item.replace(',', '\n').splitlines())
    device_ids = []
    seen = set()
    for did in raw_ids:
        did = (did or '').strip()
        if did and did not in seen:
            seen.add(did)
            device_ids.append(did)

    if not device_ids:
        flash('No devices selected for deletion.', 'error')
        return redirect(request.referrer or url_for('device_list'))

    deleted = 0
    not_found = 0
    skipped_non_retired = 0
    for device_id in device_ids:
        device = db.get_device(device_id)
        if not device:
            not_found += 1
            continue
        if (device.get('status') or '').strip() != 'retired':
            skipped_non_retired += 1
            continue

        for att in db.get_device_attachments(device_id):
            filepath = os.path.join(DEVICE_UPLOADS_DIR, str(device_id), att['filename'])
            if os.path.isfile(filepath):
                try:
                    os.remove(filepath)
                except Exception:
                    app_logger.warning('Bulk delete: failed removing attachment file %s', filepath)
        upload_dir = os.path.join(DEVICE_UPLOADS_DIR, str(device_id))
        if os.path.isdir(upload_dir):
            try:
                shutil.rmtree(upload_dir)
            except Exception:
                # Fall back to manual cleanup for transient filesystem issues.
                try:
                    for root, dirs, files in os.walk(upload_dir, topdown=False):
                        for name in files:
                            os.remove(os.path.join(root, name))
                        for name in dirs:
                            os.rmdir(os.path.join(root, name))
                    os.rmdir(upload_dir)
                except Exception:
                    app_logger.warning('Bulk delete: failed removing upload dir %s', upload_dir)

        for ext in ('png', 'pdf'):
            path = os.path.join(app.static_folder, 'labels', f'{device_id}.{ext}')
            if os.path.isfile(path):
                try:
                    os.remove(path)
                except Exception:
                    app_logger.warning('Bulk delete: failed removing label file %s', path)

        if db.delete_device(device_id):
            deleted += 1
            app_logger.info('Device deleted: id=%s name="%s" by=%s',
                            device_id, device.get('name', ''), current_username())
        else:
            not_found += 1

    if deleted:
        flash(f'Deleted {deleted} retired device(s).', 'success')
    if skipped_non_retired:
        flash(
            f'{skipped_non_retired} selected device(s) were skipped '
            f'(bulk delete only removes retired devices).',
            'warning',
        )
    if not_found:
        flash(f'{not_found} selected device(s) were not found.', 'warning')
    return redirect(request.referrer or url_for('device_list'))

# ---------------------------------------------------------------------------
# Check out / Check in (admin only)
# ---------------------------------------------------------------------------

@app.route('/devices/<device_id>/checkout', methods=['POST'])
@permission_required('devices')
def device_checkout(device_id):
    assigned_to = request.form.get('assigned_to', '').strip()
    if not assigned_to:
        flash('Please enter who is checking out this device.', 'error')
        return redirect(url_for('device_detail', device_id=device_id))

    db.checkout_device(device_id, assigned_to, performed_by=current_username())
    app_logger.info('Device checked out: id=%s to=%s by=%s', device_id, assigned_to, current_username())
    flash(f'Device checked out to {assigned_to}.', 'success')
    return redirect(url_for('device_detail', device_id=device_id))


@app.route('/devices/<device_id>/checkin', methods=['POST'])
@permission_required('devices')
def device_checkin(device_id):
    db.checkin_device(device_id, performed_by=current_username())
    app_logger.info('Device checked in: id=%s by=%s', device_id, current_username())
    flash('Device checked in successfully.', 'success')
    return redirect(url_for('device_detail', device_id=device_id))

# ---------------------------------------------------------------------------
# Label serving and label sheet generation
# ---------------------------------------------------------------------------

def _label_name(device):
    """Return the name to display on a device label. Printers use codename + variant only."""
    if device.get('category') == 'Printer' and device.get('codename'):
        variant = device.get('variant', '')
        return f"{device['codename']} {variant}".strip() if variant else device['codename']
    return device['name']


@app.route('/labels/<device_id>.png')
def serve_label(device_id):
    """Serve a label PNG, regenerating only when missing or stale."""
    device = db.get_device(device_id)
    if not device:
        app_logger.warning('Label requested for unknown device: id=%s', device_id)
        return 'Device not found', 404
    path = barcode_utils.get_label_path(device_id)
    if not os.path.isfile(path) or os.path.getmtime(path) < datetime.fromisoformat(device['updated_at']).timestamp():
        barcode_utils.generate_label(device_id, device['barcode_value'], _label_name(device))
    return send_file(path, mimetype='image/png')


@app.route('/labels/<device_id>.pdf')
def serve_label_pdf(device_id):
    """Serve a label as a PDF matching 3.5x1.5 inch landscape label stock."""
    device = db.get_device(device_id)
    if not device:
        return 'Device not found', 404
    # Regenerate only if label is missing or stale
    path = barcode_utils.get_label_path(device_id)
    if not os.path.isfile(path) or os.path.getmtime(path) < datetime.fromisoformat(device['updated_at']).timestamp():
        barcode_utils.generate_label(device_id, device['barcode_value'], _label_name(device))

    # Landscape PNG (1050x450 = 3.5x1.5" at 300 DPI)
    import zlib
    from PIL import Image
    img = Image.open(path).convert('RGB')
    img_w, img_h = img.size
    # Use FlateDecode (lossless) instead of JPEG to preserve crisp barcode edges
    raw_data = img.tobytes()
    img_data = zlib.compress(raw_data, 9)

    # Landscape page matching label stock: 3.5" wide x 1.5" tall
    page_w = 252   # 3.5 * 72
    page_h = 108   # 1.5 * 72

    xref_offsets = []
    pdf = io.BytesIO()

    pdf.write(b'%PDF-1.4\n')

    xref_offsets.append(pdf.tell())
    pdf.write(b'1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n')

    xref_offsets.append(pdf.tell())
    pdf.write(b'2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n')

    xref_offsets.append(pdf.tell())
    pdf.write(f'3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] /Contents 5 0 R /Resources << /XObject << /Img 4 0 R >> >> >>\nendobj\n'.encode())

    xref_offsets.append(pdf.tell())
    pdf.write(f'4 0 obj\n<< /Type /XObject /Subtype /Image /Width {img_w} /Height {img_h} /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length {len(img_data)} >>\nstream\n'.encode())
    pdf.write(img_data)
    pdf.write(b'\nendstream\nendobj\n')

    # Simple scale — landscape image on landscape page, no rotation needed
    content = f'q {page_w} 0 0 {page_h} 0 0 cm /Img Do Q'.encode()
    xref_offsets.append(pdf.tell())
    pdf.write(f'5 0 obj\n<< /Length {len(content)} >>\nstream\n'.encode())
    pdf.write(content)
    pdf.write(b'\nendstream\nendobj\n')

    # Xref table
    xref_start = pdf.tell()
    pdf.write(b'xref\n')
    pdf.write(f'0 {len(xref_offsets) + 1}\n'.encode())
    pdf.write(b'0000000000 65535 f \n')
    for offset in xref_offsets:
        pdf.write(f'{offset:010d} 00000 n \n'.encode())

    # Trailer
    pdf.write(f'trailer\n<< /Size {len(xref_offsets) + 1} /Root 1 0 R >>\n'.encode())
    pdf.write(b'startxref\n')
    pdf.write(f'{xref_start}\n'.encode())
    pdf.write(b'%%EOF\n')

    pdf.seek(0)
    return send_file(pdf, mimetype='application/pdf',
                     download_name=f'{device_id}_label.pdf')


@app.route('/labels/sheet', methods=['POST'])
@permission_required('devices')
def label_sheet():
    """Generate and download a printable sheet of labels for selected devices."""
    device_ids = request.form.getlist('device_ids')
    if not device_ids:
        flash('No devices selected for label printing.', 'error')
        return redirect(url_for('device_list'))

    devices = []
    for did in device_ids:
        d = db.get_device(did)
        if d:
            devices.append(d)

    if not devices:
        flash('No valid devices found.', 'error')
        return redirect(url_for('device_list'))

    sheet = barcode_utils.generate_label_sheet(devices)
    buffer = io.BytesIO()
    sheet.save(buffer, format='PNG')
    buffer.seek(0)

    app_logger.info('Label sheet generated: %d devices by=%s', len(devices), current_username())
    return send_file(buffer, mimetype='image/png', as_attachment=True,
                     download_name='label_sheet.png')

# ---------------------------------------------------------------------------
# Scanner page and API
# ---------------------------------------------------------------------------

@app.route('/scan')
def scan_page():
    app_logger.debug('Scan page accessed: ip=%s', request.remote_addr)
    activity = db.get_audit_log(limit=15)
    return render_template('scan.html', recent_activity=activity)


@app.route('/api/lookup')
def api_lookup():
    """JSON API for barcode scanner lookup. Case-insensitive."""
    barcode = request.args.get('barcode', '').strip()
    if not barcode:
        app_logger.warning('Barcode lookup: empty barcode ip=%s', request.remote_addr)
        return jsonify({'found': False, 'error': 'No barcode provided'}), 400

    device = db.get_device_by_barcode(barcode)
    if device:
        app_logger.info('Barcode scan: barcode=%s found="%s" (id=%s) ip=%s', barcode, device['name'], device['device_id'], request.remote_addr)
        return jsonify({
            'found': True,
            'device_id': device['device_id'],
            'name': device['name'],
            'status': device['status'],
            'assigned_to': device['assigned_to'],
            'location': device['location'],
        })
    else:
        app_logger.info('Barcode scan: barcode=%s not_found ip=%s', barcode, request.remote_addr)
        return jsonify({'found': False}), 404

# ---------------------------------------------------------------------------
# Export (public) — CSV and Excel
# ---------------------------------------------------------------------------

def _get_export_devices():
    """Gather devices based on export filter query params."""
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    connectivity = request.args.get('connectivity', '')
    location = request.args.get('location', '')
    q = request.args.get('q', '')
    include_retired = request.args.get('include_retired') == '1'

    if category or status or connectivity or location or q:
        if include_retired and not status:
            non_retired = db.search_devices(query=q, category=category, connectivity=connectivity, location=location)
            retired = db.search_devices(query=q, category=category, status='retired', connectivity=connectivity, location=location)
            seen = set()
            devices = []
            for d in non_retired + retired:
                if d['device_id'] not in seen:
                    seen.add(d['device_id'])
                    devices.append(d)
        else:
            devices = db.search_devices(
                query=q, category=category,
                status=status if status else '',
                connectivity=connectivity, location=location,
            )
    else:
        devices = db.get_all_devices(include_retired=include_retired)
    return devices

EXPORT_FIELDS = ['device_id', 'barcode_value', 'name', 'category', 'device_type',
                 'is_mesh', 'manufacturer',
                 'model_number', 'serial_number', 'connectivity', 'vendor_supplied',
                 'status', 'location', 'assigned_to', 'notes', 'codename', 'variant', 'hw_version',
                 'created_at', 'updated_at']

EXPORT_HEADERS = {
    'device_id': 'Device ID',
    'barcode_value': 'Barcode',
    'name': 'Name',
    'category': 'Category',
    'device_type': 'Device Type',
    'is_mesh': 'Mesh',
    'manufacturer': 'Manufacturer',
    'model_number': 'Model Number',
    'hw_version': 'HW Version',
    'serial_number': 'Serial Number',
    'connectivity': 'Connectivity Type/Version',
    'vendor_supplied': 'Source',
    'status': 'Status',
    'location': 'Location',
    'assigned_to': 'Assigned To',
    'notes': 'Notes',
    'codename': 'Codename',
    'variant': 'Variant',
    'created_at': 'Created',
    'updated_at': 'Updated',
}

# Map export column titles back to internal field names (Excel import).
IMPORT_HEADER_TO_FIELD = {label: key for key, label in EXPORT_HEADERS.items()}
# Case-insensitive / extra-whitespace header match (e.g. "category", "  Category  ")
_IMPORT_HEADER_NORMALIZED_TO_FIELD = {}
for _k, _lbl in EXPORT_HEADERS.items():
    _norm = ' '.join(_lbl.split()).casefold()
    if _norm:
        _IMPORT_HEADER_NORMALIZED_TO_FIELD[_norm] = _k

# Recognized alternatives when row-1 headers are localized or renamed.
_IMPORT_HEADER_FIELD_ALIASES = {
    '类别': 'category',
    '分類': 'category',
    'device category': 'category',
    'equipment category': 'category',
}

IMPORT_SKIP_FIELDS = frozenset({'device_id', 'created_at', 'updated_at'})
IMPORT_MAX_ROWS = 2000
_IMPORT_ALLOWED_STATUS = frozenset({'available', 'checked_out', 'retired', 'lost'})
_CUSTOM_BARCODE_ALLOWED_RE = re.compile(r'^[A-Za-z0-9-]+$')


def _normalize_custom_barcode(raw):
    """Normalize optional user-provided barcode. Empty string means auto-generate."""
    return _xlsx_cell_str(raw).strip().upper()


def _validate_custom_barcode_or_error(val):
    """Return None if valid/empty, or an error string."""
    if not val:
        return None
    if not _CUSTOM_BARCODE_ALLOWED_RE.match(val):
        return 'Custom barcode may contain only letters, numbers, and hyphen (-).'
    return None


def _import_header_to_field(label):
    """Resolve spreadsheet column title to internal field name."""
    s = _xlsx_cell_str(label)
    if not s:
        return None
    f = IMPORT_HEADER_TO_FIELD.get(s)
    if f:
        return f
    norm = ' '.join(s.split()).casefold()
    f = _IMPORT_HEADER_NORMALIZED_TO_FIELD.get(norm)
    if f:
        return f
    return _IMPORT_HEADER_FIELD_ALIASES.get(norm)


def _xlsx_raw_value_nonempty(val):
    """True if an openpyxl cell value should be treated as present (not blank)."""
    if val is None:
        return False
    if isinstance(val, str):
        return val.strip() != ''
    return True


def _xlsx_ws_cell_effective_value(ws, row_idx, col_idx):
    """Raw cell value from a worksheet, using merged range top-left when the cell is blank."""
    cell = ws.cell(row=row_idx, column=col_idx)
    v = cell.value
    if _xlsx_raw_value_nonempty(v):
        return v
    for mrange in ws.merged_cells.ranges:
        if mrange.min_row <= row_idx <= mrange.max_row and mrange.min_col <= col_idx <= mrange.max_col:
            tl = ws.cell(row=mrange.min_row, column=mrange.min_col).value
            if _xlsx_raw_value_nonempty(tl):
                return tl
            # Some exports keep text on the first row of a merge but not at top-left (multi-column).
            for cc in range(mrange.min_col, mrange.max_col + 1):
                alt = ws.cell(row=mrange.min_row, column=cc).value
                if _xlsx_raw_value_nonempty(alt):
                    return alt
    return v


def _xlsx_pick_worksheet_for_device_import(wb):
    """Prefer the Inventory sheet; otherwise pick the sheet whose row 1 maps to the most known headers."""
    by_title = {s.title.casefold(): s for s in wb.worksheets}
    if 'inventory' in by_title:
        return by_title['inventory']
    best_ws, best_score = None, -1
    for cand in wb.worksheets:
        max_col = cand.max_column or 0
        if max_col < 1:
            continue
        score = 0
        for col_idx in range(1, max_col + 1):
            hv = _xlsx_ws_cell_effective_value(cand, 1, col_idx)
            field = _import_header_to_field(hv)
            if field and field not in IMPORT_SKIP_FIELDS:
                score += 1
        if score > best_score:
            best_score = score
            best_ws = cand
    return best_ws if best_ws is not None else wb.active


def _xlsx_cell_str(val):
    """Normalize an openpyxl cell value to a string for import."""
    if val is None:
        return ''
    if isinstance(val, bool):
        return 'Yes' if val else 'No'
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val).strip()


def _parse_vendor_supplied_import(val):
    s = _xlsx_cell_str(val).lower()
    if s in ('1', 'true', 'yes', 'vendor supplied', 'y'):
        return 1
    return 0


def _parse_is_mesh_import(val):
    s = _xlsx_cell_str(val).lower()
    return 1 if s in ('1', 'true', 'yes', 'y') else 0


def _parse_status_import(val):
    """Return (status or None, error_message or None)."""
    s = _xlsx_cell_str(val).lower().replace(' ', '_').replace('-', '_')
    if not s:
        return 'available', None
    if s == 'checkedout':
        s = 'checked_out'
    if s in _IMPORT_ALLOWED_STATUS:
        return s, None
    return None, f'invalid status {val!r} (use: available, checked out, retired, lost)'


def _import_row_seems_blank(clean_row):
    keys = ('category', 'manufacturer', 'model_number', 'serial_number', 'name', 'codename')
    return not any((clean_row.get(k) or '').strip() for k in keys)


def _import_serial_normalized(raw):
    """Treat common spreadsheet placeholders as no serial (avoids duplicate literal '-')."""
    s = _xlsx_cell_str(raw)
    if not s:
        return ''
    # Map unicode dashes / minus signs to ASCII hyphen for token checks
    dash_norm = s.translate(str.maketrans('\u2012\u2013\u2014\u2015\u2212', '-----'))
    # Strip all whitespace (incl. NBSP / ZWSP from Excel paste) for placeholder match only
    token = ''.join(ch for ch in dash_norm if not ch.isspace()).casefold()
    if token in ('-', '--', 'n/a', 'na', 'none', '.', 'tbd', 'tba', 'unknown', '?', 'null', 'nil'):
        return ''
    return s


def _device_data_from_import_row(clean_row):
    """Validate a mapped row dict (string values) like device_add. Returns (data dict or None, errors list)."""
    manufacturer = (clean_row.get('manufacturer') or '').strip()
    model_number = (clean_row.get('model_number') or '').strip()
    category = (clean_row.get('category') or '').strip()
    codename = (clean_row.get('codename') or '').strip()
    if codename.upper() == 'N/A':
        codename = ''
    serial_number = _import_serial_normalized(clean_row.get('serial_number'))
    custom_barcode = _normalize_custom_barcode(clean_row.get('barcode_value', ''))
    variant_raw = (clean_row.get('variant') or '').strip()
    if variant_raw.upper() == 'N/A':
        variant_raw = ''

    errors = []
    if category == 'Printer':
        if not manufacturer:
            manufacturer = 'HP'
        if not codename:
            errors.append('Codename is required for printers.')
    elif category in ('Connectivity Device', 'Endpoint Device'):
        if not manufacturer:
            errors.append('Manufacturer is required.')
        if not model_number:
            errors.append('Model number is required.')

    if errors:
        return None, errors
    bc_err = _validate_custom_barcode_or_error(custom_barcode)
    if bc_err:
        return None, [bc_err]

    if category == 'Printer':
        variant = variant_raw
        codename_display = f'{codename} {variant}'.strip() if variant else codename
        mfg_model = f'{manufacturer} {model_number}'.strip()
        name = f'{codename_display} ({mfg_model})' if mfg_model else codename_display
    else:
        name = f'{manufacturer} {model_number}'.strip()
        if not name:
            name = (clean_row.get('name') or '').strip()
        if not name:
            return None, ['Name or manufacturer and model number are required.']

    vendor_supplied = _parse_vendor_supplied_import(clean_row.get('vendor_supplied'))
    if category != 'Connectivity Device':
        vendor_supplied = 0

    status, st_err = _parse_status_import(clean_row.get('status'))
    if st_err:
        return None, [st_err]

    device_type = (clean_row.get('device_type') or '').strip()
    if device_type.upper() == 'N/A':
        device_type = ''

    data = {
        'name': name,
        'category': category,
        'barcode_value': custom_barcode,
        'manufacturer': manufacturer,
        'model_number': model_number,
        'hw_version': (clean_row.get('hw_version') or '').strip(),
        'serial_number': serial_number,
        'connectivity': (clean_row.get('connectivity') or '').strip(),
        'vendor_supplied': vendor_supplied,
        'location': (clean_row.get('location') or '').strip(),
        'notes': (clean_row.get('notes') or '').strip(),
        'codename': codename or ('N/A' if category != 'Printer' else ''),
        'variant': variant_raw or ('N/A' if category != 'Printer' else ''),
        'device_type': device_type or ('N/A' if category not in ('Connectivity Device', 'Endpoint Device') else ''),
        'is_mesh': _parse_is_mesh_import(clean_row.get('is_mesh')),
        'status': status,
        'assigned_to': (clean_row.get('assigned_to') or '').strip(),
    }
    return data, []

@app.route('/export')
def export_csv():
    """Export devices to CSV."""
    devices = _get_export_devices()
    app_logger.info('CSV export: %d devices ip=%s', len(devices), request.remote_addr)

    output = io.StringIO()
    headers = [EXPORT_HEADERS.get(f, f) for f in EXPORT_FIELDS]
    writer = csv.writer(output)
    writer.writerow(headers)
    for d in devices:
        row = []
        for f in EXPORT_FIELDS:
            val = d.get(f, '')
            if f == 'vendor_supplied':
                val = 'Vendor Supplied' if val else 'HP Owned'
            elif f == 'is_mesh':
                val = 'Yes' if val else 'No'
            row.append(val if val is not None else '')
        writer.writerow(row)

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=inventory_export.csv'}
    )


@app.route('/export/xlsx')
def export_xlsx():
    """Export devices to Excel (.xlsx)."""
    devices = _get_export_devices()
    app_logger.info('Excel export: %d devices ip=%s', len(devices), request.remote_addr)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required for Excel export. Install with: pip install openpyxl', 'error')
        return redirect(url_for('device_list'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # Header row
    headers = [EXPORT_HEADERS.get(f, f) for f in EXPORT_FIELDS]
    ws.append(headers)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for d in devices:
        row = []
        for f in EXPORT_FIELDS:
            val = d.get(f, '')
            if f == 'vendor_supplied':
                val = 'Vendor Supplied' if val else 'HP Owned'
            elif f == 'is_mesh':
                val = 'Yes' if val else 'No'
            row.append(val if val is not None else '')
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=inventory_export.xlsx'}
    )


@app.route('/devices/import/xlsx', methods=['POST'])
@permission_required('devices')
def import_devices_xlsx():
    """Import devices from an Excel file (.xlsx). Columns must match Export → Excel."""
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl is required for Excel import.', 'error')
        return redirect(url_for('device_list'))

    upload = request.files.get('file')
    if not upload or not upload.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('device_list'))
    if not upload.filename.lower().endswith('.xlsx'):
        flash('Please upload an .xlsx file.', 'error')
        return redirect(url_for('device_list'))

    try:
        upload.seek(0)
    except (AttributeError, OSError, IOError):
        pass

    try:
        xlsx_bytes = upload.read()
    except Exception as e:
        app_logger.warning('Excel import read upload failed: %s', e)
        flash(f'Could not read uploaded file: {e}', 'error')
        return redirect(url_for('device_list'))

    if not xlsx_bytes:
        flash('The uploaded file is empty.', 'error')
        return redirect(url_for('device_list'))

    wb = None
    try:
        try:
            # data_only=False: merged / label cells often have no cached value with data_only=True
            # in files saved from Excel, which makes Category (and others) read as blank.
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=False, data_only=False)
        except Exception as e:
            app_logger.warning('Excel import read failed: %s', e)
            flash(f'Could not read Excel file: {e}', 'error')
            return redirect(url_for('device_list'))

        ws = _xlsx_pick_worksheet_for_device_import(wb)
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 1 or max_col < 1:
            flash('The spreadsheet is empty.', 'error')
            return redirect(url_for('device_list'))

        header_fields = []
        for col_idx in range(1, max_col + 1):
            hv = _xlsx_ws_cell_effective_value(ws, 1, col_idx)
            header_fields.append(_import_header_to_field(hv))

        if not any(f for f in header_fields if f and f not in IMPORT_SKIP_FIELDS):
            flash('No recognized column headers. Use Export → Excel as a template.', 'error')
            return redirect(url_for('device_list'))

        if 'category' not in header_fields:
            flash(
                'No Category column found in row 1. Use Export → Excel as a template, '
                'keep the header exactly "Category" (or 类别), and do not delete that column.',
                'error',
            )
            return redirect(url_for('device_list'))

        imported = 0
        err_lines = []
        data_row_count = 0
        last_import_category = ''
        scan_limit = min(max_row, max(IMPORT_MAX_ROWS * 5, 5000))

        for row_idx in range(2, scan_limit + 1):
            raw = {}
            for col_idx in range(1, max_col + 1):
                field = header_fields[col_idx - 1]
                if not field or field in IMPORT_SKIP_FIELDS:
                    continue
                val = _xlsx_ws_cell_effective_value(ws, row_idx, col_idx)
                if field in raw:
                    prev = raw[field]
                    if _xlsx_raw_value_nonempty(prev) and not _xlsx_raw_value_nonempty(val):
                        continue
                raw[field] = val

            clean = {k: _xlsx_cell_str(v) for k, v in raw.items()}

            if _import_row_seems_blank(clean):
                continue

            data_row_count += 1
            if data_row_count > IMPORT_MAX_ROWS:
                err_lines.append(f'Row {row_idx}: skipped (max {IMPORT_MAX_ROWS} data rows)')
                break

            cat_cell = (clean.get('category') or '').strip()
            if not cat_cell and last_import_category:
                clean = dict(clean)
                clean['category'] = last_import_category
            elif cat_cell:
                last_import_category = cat_cell

            if not (clean.get('category') or '').strip():
                err_lines.append(
                    f'Row {row_idx}: Category is required '
                    f'(fill Category on the first row of each group, or unmerge the Category column).'
                )
                continue

            data, v_errors = _device_data_from_import_row(clean)
            if v_errors:
                for e in v_errors:
                    err_lines.append(f'Row {row_idx}: {e}')
                continue

            if data.get('serial_number'):
                existing = db.get_device_by_serial(data['serial_number'])
                if existing:
                    err_lines.append(
                        f'Row {row_idx}: Serial "{data["serial_number"]}" already exists ({existing["name"]}).'
                    )
                    continue
            if data.get('barcode_value'):
                existing_bc = db.get_device_by_barcode(data['barcode_value'])
                if existing_bc:
                    err_lines.append(
                        f'Row {row_idx}: Barcode "{data["barcode_value"]}" already exists ({existing_bc["name"]}).'
                    )
                    continue

            try:
                device_id = db.add_device(data, performed_by=current_username())
                device = db.get_device(device_id)
                barcode_utils.generate_label(device_id, device['barcode_value'], _label_name(device))
                imported += 1
            except Exception as e:
                app_logger.exception('Excel import row failed at row %s', row_idx)
                err_lines.append(f'Row {row_idx}: {e}')

        app_logger.info(
            'Excel import: imported=%d errors=%d rows=%d by=%s ip=%s',
            imported, len(err_lines), data_row_count, current_username(), request.remote_addr,
        )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    if imported:
        flash(f'Successfully imported {imported} device(s).', 'success')
    if err_lines:
        cap = 25
        tail = err_lines[:cap]
        msg = 'Import issues: ' + ' | '.join(tail)
        if len(err_lines) > cap:
            msg += f' …and {len(err_lines) - cap} more.'
        flash(msg, 'warning' if imported else 'error')

    if not imported and not err_lines:
        flash('No data rows were imported (file had no non-empty rows).', 'warning')

    return redirect(url_for('device_list'))


# ---------------------------------------------------------------------------
# User management (admin only)
# ---------------------------------------------------------------------------

@app.route('/users')
@permission_required('users')
def user_list():
    users = db.get_all_users()
    guest_permissions = db.get_guest_permissions()
    return render_template('users.html', users=users,
                           guest_permissions=guest_permissions,
                           guest_assignable_permissions=GUEST_ASSIGNABLE_PERMISSIONS)


@app.route('/users/add', methods=['GET', 'POST'])
@permission_required('users')
def user_add():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        role = request.form.get('role', 'custom')
        display_name = request.form.get('display_name', '').strip()
        password_hint = request.form.get('password_hint', '').strip()

        # Collect permissions from checkboxes (only for custom role)
        permissions = None
        if role == 'custom':
            permissions = request.form.getlist('permissions')

        if not username or not password:
            flash('Username and password are required.', 'error')
            return render_template('user_form.html', user={}, is_edit=False,
                                   assignable_permissions=ASSIGNABLE_PERMISSIONS)

        if len(password) < 4:
            flash('Password must be at least 4 characters.', 'error')
            return render_template('user_form.html', user=request.form, is_edit=False,
                                   assignable_permissions=ASSIGNABLE_PERMISSIONS)

        try:
            db.create_user(username, password, role, display_name, permissions=permissions,
                           password_hint=password_hint)
            app_logger.info('User created: username=%s role=%s permissions=%s by=%s',
                            username, role, permissions, current_username())
            flash(f'User "{username}" created successfully.', 'success')
            return redirect(url_for('user_list'))
        except ValueError as e:
            flash(str(e), 'error')
            return render_template('user_form.html', user=request.form, is_edit=False,
                                   assignable_permissions=ASSIGNABLE_PERMISSIONS)

    return render_template('user_form.html', user={}, is_edit=False,
                           assignable_permissions=ASSIGNABLE_PERMISSIONS)


@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@permission_required('users')
def user_edit(user_id):
    user = db.get_user(user_id)
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('user_list'))

    if request.method == 'POST':
        role = request.form.get('role', user['role'])
        data = {
            'display_name': request.form.get('display_name', '').strip(),
            'role': role,
            'password_hint': request.form.get('password_hint', '').strip(),
        }
        if role == 'custom':
            data['permissions'] = request.form.getlist('permissions')
        else:
            data['permissions'] = None  # admin uses role defaults

        password = request.form.get('password', '').strip()
        if password:
            if len(password) < 4:
                flash('Password must be at least 4 characters.', 'error')
                return render_template('user_form.html', user=user, is_edit=True,
                                       assignable_permissions=ASSIGNABLE_PERMISSIONS)
            data['password'] = password

        db.update_user(user_id, data)
        app_logger.info('User updated: username=%s role=%s by=%s',
                        user['username'], role, current_username())
        flash(f'User "{user["username"]}" updated.', 'success')
        return redirect(url_for('user_list'))

    return render_template('user_form.html', user=user, is_edit=True,
                           assignable_permissions=ASSIGNABLE_PERMISSIONS)


@app.route('/users/<int:user_id>/delete', methods=['POST'])
@permission_required('users')
def user_delete(user_id):
    try:
        db.delete_user(user_id)
        app_logger.info('User deleted: user_id=%s by=%s', user_id, current_username())
        flash('User deleted.', 'success')
    except ValueError as e:
        app_logger.warning('User delete failed: user_id=%s error=%s', user_id, e)
        flash(str(e), 'error')
    return redirect(url_for('user_list'))

# ---------------------------------------------------------------------------
# Application Log viewer (admin only)
# ---------------------------------------------------------------------------

@app.route('/logs')
@permission_required('logs')
def app_logs():
    """View application log entries with pagination. Most recent first."""
    per_page = 200
    page = max(1, request.args.get('page', 1, type=int))

    lines = []
    # Read rotated backup first (older), then current log (newer)
    for log_path in [LOG_FILE + '.1', LOG_FILE]:
        try:
            with open(log_path, 'r') as f:
                lines.extend(f.readlines())
        except FileNotFoundError:
            pass

    # Parse into structured entries, most recent first
    all_entries = []
    for line in reversed(lines):
        line = line.strip()
        if not line:
            continue
        parts = line.split(' | ', 2)
        if len(parts) == 3:
            all_entries.append({
                'timestamp': parts[0],
                'level': parts[1].strip(),
                'message': parts[2],
            })
        else:
            all_entries.append({
                'timestamp': '',
                'level': '',
                'message': line,
            })

    total = len(all_entries)
    total_pages = max(1, (total + per_page - 1) // per_page)
    start = (page - 1) * per_page
    entries = all_entries[start:start + per_page]

    log_config = _load_log_config()
    log_file_size = sum(os.path.getsize(p) for p in [LOG_FILE, LOG_FILE + '.1'] if os.path.exists(p))
    return render_template('app_log.html', entries=entries, log_config=log_config,
                           log_file_size=log_file_size, page=page, total_pages=total_pages)


@app.route('/logs/clear', methods=['POST'])
@permission_required('logs')
def clear_logs():
    """Clear the application log file."""
    try:
        with open(LOG_FILE, 'w') as f:
            f.write('')
        # Remove rotated backup file too
        backup_log = LOG_FILE + '.1'
        if os.path.exists(backup_log):
            os.remove(backup_log)
        app_logger.info('Application log cleared by %s', current_username())
        flash('Application log cleared.', 'success')
    except Exception as e:
        app_logger.error('Log clear failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Error clearing log: {e}', 'error')
    return redirect(url_for('app_logs'))


@app.route('/logs/export')
@permission_required('logs')
def export_logs():
    """Export the application log as a downloadable .log file."""
    lines = []
    for log_path in [LOG_FILE + '.1', LOG_FILE]:
        try:
            with open(log_path, 'r') as f:
                lines.extend(f.readlines())
        except FileNotFoundError:
            pass

    content = ''.join(lines)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return Response(
        content,
        mimetype='text/plain',
        headers={'Content-Disposition': f'attachment; filename=app_log_{timestamp}.log'}
    )


@app.route('/logs/config', methods=['POST'])
@permission_required('logs')
def update_log_config():
    """Update application log max size."""
    try:
        max_size_mb = float(request.form.get('max_size_mb', 2))
        if max_size_mb < 0.1:
            max_size_mb = 0.1
        if max_size_mb > 100:
            max_size_mb = 100
    except (ValueError, TypeError):
        max_size_mb = 2

    config = {'max_size_mb': max_size_mb}
    _save_log_config(config)
    _reconfigure_log_handler(max_size_mb)
    app_logger.info('Log max size changed to %.1f MB by %s', max_size_mb, current_username())
    flash(f'Log max size set to {max_size_mb} MB. Log will overwrite oldest entries when this limit is reached.', 'success')
    return redirect(url_for('app_logs'))

# ---------------------------------------------------------------------------
# Change own password (any logged-in user)
# ---------------------------------------------------------------------------

@app.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    def _render(**extra):
        users = db.get_all_users() if has_permission('users') else []
        server_config = _load_server_config()
        autostart = {
            'supported': sys.platform == 'win32',
            'enabled': _get_autostart_enabled(),
        }
        return render_template('account.html', users=users, server_config=server_config,
                               autostart=autostart, **extra)

    if request.method == 'POST':
        # Password change + hint form
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        hint = request.form.get('password_hint', '').strip()

        # Verify current password
        user = db.authenticate_user(g.user['username'], current_pw)
        if not user:
            flash('Current password is incorrect.', 'error')
            return _render()

        if len(new_pw) < 4:
            flash('New password must be at least 4 characters.', 'error')
            return _render()

        if new_pw != confirm_pw:
            flash('New passwords do not match.', 'error')
            return _render()

        db.update_user(g.user['user_id'], {'password': new_pw, 'password_hint': hint})
        app_logger.info('Password changed: user=%s', g.user['username'])
        flash('Password changed successfully.', 'success')
        return redirect(url_for('account'))

    return _render()


@app.route('/docs')
@login_required
def docs():
    """Render the README as in-app documentation."""
    # Check multiple locations — PyInstaller _MEIPASS root, next to app.py,
    # and DATA_DIR (for dev). Does NOT depend on git or internet access.
    candidates = [
        os.path.join(BUNDLE_DIR, 'README.md'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'README.md'),
        os.path.join(DATA_DIR, 'README.md'),
    ]
    content = None
    for readme_path in candidates:
        try:
            with open(readme_path, 'r', encoding='utf-8') as f:
                content = f.read()
                break
        except (FileNotFoundError, OSError):
            continue
    if content is None:
        content = '*Documentation file not found.*'
    return render_template('docs.html', readme_content=content)


@app.route('/settings/server', methods=['POST'])
@permission_required('settings')
def save_server_config():
    try:
        port = int(request.form.get('port', 8080))
        if port < 1 or port > 65535:
            flash('Port must be between 1 and 65535.', 'error')
            return redirect(url_for('account'))
    except (ValueError, TypeError):
        flash('Invalid port number.', 'error')
        return redirect(url_for('account'))

    config = _load_server_config()
    config['port'] = port

    # Update repo URL & branch (for the self-update feature)
    # Store empty string when using the default — keeps the admin UI blank
    # so users only see their custom override (if any).
    repo_url = request.form.get('update_repo_url', '').strip()
    if repo_url == _DEFAULT_UPDATE_REPO_URL:
        repo_url = ''
    config['update_repo_url'] = repo_url
    config['update_branch'] = request.form.get('update_branch', '').strip()

    _save_server_config(config)
    app_logger.info('Server config updated: port=%d repo=%s branch=%s by user=%s',
                    port, repo_url or '(default)', config['update_branch'] or '(releases)',
                    g.user['username'])

    # Handle Windows autostart toggle
    autostart_enabled = '1' in request.form.getlist('autostart_enabled')
    if sys.platform == 'win32':
        current = _get_autostart_enabled()
        if current != autostart_enabled:
            ok, msg = _set_autostart(autostart_enabled)
            app_logger.info('Autostart %s by user=%s (%s)',
                            'enabled' if autostart_enabled else 'disabled',
                            g.user['username'], msg)
            if ok:
                flash(msg, 'success')
            else:
                flash(msg, 'error')

    flash('Server settings saved. Restart the application for port changes to take effect.', 'success')
    return redirect(url_for('account'))


@app.route('/settings/guest-permissions', methods=['POST'])
@permission_required('settings')
def save_guest_permissions():
    """Save which permissions are granted to non-logged-in (guest) users."""
    # Only allow permissions from the guest-assignable list
    allowed_keys = {k for k, _ in GUEST_ASSIGNABLE_PERMISSIONS}
    selected = set(request.form.getlist('guest_permissions'))
    # Filter to only valid permission keys
    valid = selected & allowed_keys
    db.save_guest_permissions(valid)
    app_logger.info('Guest permissions updated to %s by user=%s', sorted(valid), g.user['username'])
    flash('Public access permissions saved.', 'success')
    return redirect(url_for('account'))


# ---------------------------------------------------------------------------
# Application self-update (admin only)
# ---------------------------------------------------------------------------

_REPO_DIR = os.path.dirname(os.path.abspath(__file__))


def _parse_version(v):
    """Parse 'X.Y.Z' into tuple of ints for comparison."""
    try:
        return tuple(int(x) for x in v.lstrip('v').split('.'))
    except (ValueError, AttributeError):
        return (0,)


def _parse_github_repo(url):
    """Extract (owner, repo) from a GitHub URL. Returns (None, None) for
    non-GitHub URLs."""
    import re
    m = re.match(r'(?:https?://)?(?:www\.)?github\.com/([^/]+)/([^/]+?)(?:\.git)?/?$', url or '')
    if m:
        return m.group(1), m.group(2)
    return None, None


def _github_api_latest_release(owner, repo, timeout=10):
    """Query GitHub API for the latest release. Returns the parsed JSON
    dict or raises on error. No auth required for public repos."""
    import urllib.request
    api_url = f'https://api.github.com/repos/{owner}/{repo}/releases/latest'
    req = urllib.request.Request(api_url, headers={
        'Accept': 'application/vnd.github+json',
        'User-Agent': 'HP-Connectivity-Inventory-Updater',
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode('utf-8'))


@app.route('/admin/update/check', methods=['POST'])
@permission_required('update')
def app_update_check():
    """Check for new tagged releases on the configured remote.

    Uses the GitHub REST API for github.com repos (no git clone needed
    and works from the PyInstaller bundle where there is no local repo).
    Falls back to `git ls-remote` for non-GitHub URLs (internal git
    servers) so admins can point at a private branch.
    """
    config = _load_server_config()
    repo_url = _get_update_repo_url(config)
    branch = config.get('update_branch', '').strip()

    if not repo_url:
        return jsonify({'error': 'No update repository configured. Set one in Server Settings.'})

    owner, repo = _parse_github_repo(repo_url)

    # --- GitHub REST API path (preferred) -------------------------------
    if owner and repo and not branch:
        try:
            release = _github_api_latest_release(owner, repo)
            latest_tag = release.get('tag_name', '').strip()
            if not latest_tag:
                return jsonify({'updates_available': False,
                                'output': f'Current: v{_app_version}\nNo releases found on {owner}/{repo}.'})

            latest_version = latest_tag.lstrip('v')
            if _parse_version(latest_version) > _parse_version(_app_version):
                body = (release.get('body') or '').strip()
                # Cap the body at ~2000 chars so the UI doesn't get overwhelmed
                if len(body) > 2000:
                    body = body[:2000] + '\n...'
                output = (f'Current: v{_app_version}\n'
                          f'Available: {latest_tag}\n'
                          f'Published: {release.get("published_at", "")}\n\n'
                          f'{body}' if body else
                          f'Current: v{_app_version}\nAvailable: {latest_tag}')
                return jsonify({
                    'updates_available': True,
                    'output': output,
                    'tag': latest_tag,
                    'release_url': release.get('html_url', ''),
                })
            else:
                return jsonify({'updates_available': False,
                                'output': f'Current: v{_app_version}\n'
                                          f'Latest release: {latest_tag}\n\n'
                                          f'Application is up to date.'})
        except Exception as e:
            app_logger.warning('GitHub API update check failed: %s', e)
            return jsonify({'error': f'Could not reach GitHub API: {e}'})

    # --- git ls-remote path (for private branches or non-GitHub) --------
    try:
        if branch:
            # Check the HEAD commit of the configured branch
            ls_result = subprocess.run(
                [GIT_EXECUTABLE, 'ls-remote', '--heads', repo_url, branch],
                capture_output=True, text=True, timeout=30,
            )
            if ls_result.returncode != 0:
                return jsonify({'error': f'git ls-remote failed: {ls_result.stderr.strip()}'})
            if not ls_result.stdout.strip():
                return jsonify({'updates_available': False,
                                'output': f'Current: v{_app_version}\nBranch "{branch}" not found on remote.'})
            remote_sha = ls_result.stdout.split()[0][:12]
            return jsonify({
                'updates_available': True,
                'output': f'Current: v{_app_version}\n'
                          f'Remote branch: {branch}\n'
                          f'HEAD: {remote_sha}\n\n'
                          f'Apply will pull the latest commit from {branch}.',
                'tag': f'branch:{branch}',
            })
        else:
            # Check tags on the remote
            ls_result = subprocess.run(
                [GIT_EXECUTABLE, 'ls-remote', '--tags', '--refs', repo_url, 'v*'],
                capture_output=True, text=True, timeout=30,
            )
            if ls_result.returncode != 0:
                return jsonify({'error': f'git ls-remote failed: {ls_result.stderr.strip()}'})
            tags = []
            for line in ls_result.stdout.splitlines():
                parts = line.strip().split('refs/tags/')
                if len(parts) == 2:
                    tags.append(parts[1])
            if not tags:
                return jsonify({'updates_available': False,
                                'output': f'Current: v{_app_version}\nNo releases found.'})
            tags.sort(key=_parse_version)
            latest_tag = tags[-1]
            if _parse_version(latest_tag) > _parse_version(_app_version):
                return jsonify({
                    'updates_available': True,
                    'output': f'Current: v{_app_version}\nAvailable: {latest_tag}',
                    'tag': latest_tag,
                })
            return jsonify({'updates_available': False,
                            'output': f'Current: v{_app_version}\n'
                                      f'Latest release: {latest_tag}\n\n'
                                      f'Application is up to date.'})
    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Git operation timed out'})
    except FileNotFoundError:
        return jsonify({'error': 'Git is not available. Configure a GitHub repository to use the API instead.'})
    except Exception as e:
        return jsonify({'error': str(e)})


def _apply_frozen_update(tag, owner, repo):
    """Download the Windows release asset for `tag` and schedule a restart
    that swaps the new exe bundle in. Only called in PyInstaller frozen mode."""
    import tempfile
    import urllib.request
    import urllib.error

    # Find the Windows zip asset in the release
    release = _github_api_latest_release(owner, repo, timeout=15)
    assets = release.get('assets') or []
    asset = None
    for a in assets:
        name = (a.get('name') or '').lower()
        if name.endswith('.zip') and 'windows' in name:
            asset = a
            break
    if not asset:
        raise RuntimeError('No Windows release asset found on GitHub — cannot update.')

    download_url = asset['browser_download_url']
    asset_size = asset.get('size', 0)
    app_logger.info('Downloading update asset: %s (%d bytes)', download_url, asset_size)

    # Download to a temp file
    updates_dir = os.path.join(DATA_DIR, 'updates')
    os.makedirs(updates_dir, exist_ok=True)
    zip_path = os.path.join(updates_dir, f'InventorySystem-{tag}.zip')
    req = urllib.request.Request(download_url, headers={
        'User-Agent': 'HP-Connectivity-Inventory-Updater',
    })
    with urllib.request.urlopen(req, timeout=300) as resp, open(zip_path, 'wb') as f:
        while True:
            chunk = resp.read(65536)
            if not chunk:
                break
            f.write(chunk)
    app_logger.info('Download complete: %s', zip_path)

    # Write a Windows batch updater that:
    #   1. waits for the current exe to exit
    #   2. extracts the new zip over the install directory
    #   3. relaunches the exe
    exe_path = sys.executable
    install_dir = os.path.dirname(exe_path)
    extract_dir = os.path.join(updates_dir, f'extract-{tag}')
    batch_path = os.path.join(updates_dir, 'apply_update.bat')
    batch = f"""@echo off
echo Waiting for InventorySystem to exit...
timeout /t 3 /nobreak > nul
taskkill /f /im InventorySystem.exe > nul 2>&1
timeout /t 1 /nobreak > nul

echo Extracting update...
if exist "{extract_dir}" rd /s /q "{extract_dir}"
mkdir "{extract_dir}"
powershell -NoProfile -Command "Expand-Archive -Force -LiteralPath '{zip_path}' -DestinationPath '{extract_dir}'"

echo Copying new files...
robocopy "{extract_dir}" "{install_dir}" /E /NFL /NDL /NJH /NJS /NC /NS /NP

echo Cleaning up...
rd /s /q "{extract_dir}"
del "{zip_path}"

echo Restarting...
start "" "{exe_path}"

exit /b 0
"""
    with open(batch_path, 'w') as f:
        f.write(batch)
    app_logger.info('Update script written: %s', batch_path)

    # Launch the batch file detached and exit
    def _run_updater():
        import time as _t
        _t.sleep(2)
        app_logger.info('Launching update script and exiting...')
        try:
            if sys.platform == 'win32':
                # CREATE_NEW_PROCESS_GROUP | DETACHED_PROCESS = 0x00000200 | 0x00000008
                subprocess.Popen(['cmd', '/c', batch_path],
                                 creationflags=0x00000208,
                                 cwd=updates_dir,
                                 close_fds=True)
            else:
                subprocess.Popen(['bash', batch_path], cwd=updates_dir)
        finally:
            os._exit(0)

    threading.Thread(target=_run_updater, daemon=True).start()
    return {
        'ok': True,
        'output': f'Downloaded {tag} ({asset_size // 1024} KB). '
                  f'The application will exit and restart shortly.',
    }


@app.route('/admin/update/apply', methods=['POST'])
@permission_required('update')
def app_update_apply():
    """Apply an update. For frozen builds, downloads the release zip and
    runs an external updater script. For source installs, uses git checkout."""
    try:
        tag = request.json.get('tag') if request.is_json else request.form.get('tag')
        if not tag:
            return jsonify({'error': 'No update tag specified'})

        config = _load_server_config()
        repo_url = _get_update_repo_url(config)
        owner, repo = _parse_github_repo(repo_url)

        # --- Frozen (PyInstaller) mode: download + swap ------------------
        if getattr(sys, 'frozen', False):
            if tag.startswith('branch:'):
                return jsonify({'error': 'Branch-based updates are only supported for source installs.'})
            if not (owner and repo):
                return jsonify({'error': 'Only GitHub repositories are supported for frozen updates.'})
            result = _apply_frozen_update(tag, owner, repo)
            app_logger.info('Frozen update initiated by=%s tag=%s', current_username(), tag)
            return jsonify(result)

        # --- Source mode: git checkout the tag (or pull the branch) ------
        if tag.startswith('branch:'):
            branch = tag[len('branch:'):]
            pull_result = subprocess.run(
                [GIT_EXECUTABLE, 'pull', repo_url, branch],
                cwd=_REPO_DIR, capture_output=True, text=True, timeout=60,
            )
            if pull_result.returncode != 0:
                return jsonify({'error': f'git pull failed: {pull_result.stderr.strip()}'})
            output = f'Updated to {branch}\n{pull_result.stdout.strip()}'
        else:
            if not tag.startswith('v'):
                return jsonify({'error': 'No valid release tag specified'})
            # Ensure tags are local
            subprocess.run(
                [GIT_EXECUTABLE, 'fetch', '--tags', '--force'],
                cwd=_REPO_DIR, capture_output=True, text=True, timeout=30,
            )
            checkout_result = subprocess.run(
                [GIT_EXECUTABLE, 'checkout', tag],
                cwd=_REPO_DIR, capture_output=True, text=True, timeout=30,
            )
            if checkout_result.returncode != 0:
                return jsonify({'error': f'Git checkout failed: {checkout_result.stderr.strip()}'})
            output = f'Updated to {tag}\n{checkout_result.stdout.strip()}'

        # Install updated dependencies (best-effort)
        pip_cmd = [sys.executable, '-m', 'pip', 'install', '-r',
                   os.path.join(_REPO_DIR, 'requirements.txt'), '-q']
        pip_result = subprocess.run(
            pip_cmd, cwd=_REPO_DIR, capture_output=True, text=True, timeout=120,
        )
        if pip_result.returncode == 0:
            output += '\nDependencies updated.'
        else:
            output += f'\nDependency install note: {pip_result.stderr.strip()}'

        app_logger.info('Application update applied by=%s: %s', current_username(), output.replace('\n', ' | '))

        def _restart():
            import time as _t
            _t.sleep(1.5)
            app_logger.info('Restarting application after update...')
            os.execv(sys.executable, [sys.executable] + sys.argv)

        threading.Thread(target=_restart, daemon=True).start()
        return jsonify({'ok': True, 'output': output})

    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Operation timed out'})
    except FileNotFoundError:
        return jsonify({'error': 'Git is not installed on this system'})
    except Exception as e:
        app_logger.error('Application update failed: %s\n%s', e, traceback.format_exc())
        return jsonify({'error': str(e)})


# ---------------------------------------------------------------------------
# Database backup (admin only)
# ---------------------------------------------------------------------------

import threading

# ---------------------------------------------------------------------------
# Persistent backup scheduler — single thread that wakes every 60 seconds
# and checks what tasks are due. Replaces fragile threading.Timer chains that
# silently died when a daemon thread was killed or an exception escaped.
# ---------------------------------------------------------------------------

_scheduler_thread = None
_scheduler_stop = threading.Event()

# Next-run timestamps (None = disabled). Protected by _scheduler_lock.
_scheduler_lock = threading.Lock()
_next_backup_time = None
_next_git_push_time = None
_next_filepath_push_time = None
_next_prune_time = None
_next_verify_time = None

# Consecutive failure counters for retry backoff (max 3 retries then normal interval)
_RETRY_DELAYS_MIN = [2, 5, 15]  # minutes to wait before retry 1, 2, 3
_fail_count = {'backup': 0, 'git_push': 0, 'filepath_push': 0, 'prune': 0}


def _scheduler_loop():
    """Persistent loop: wake every 60s, run any overdue tasks."""
    while not _scheduler_stop.is_set():
        try:
            now = datetime.now()

            with _scheduler_lock:
                run_backup = _next_backup_time is not None and now >= _next_backup_time
                run_git = _next_git_push_time is not None and now >= _next_git_push_time
                run_filepath = _next_filepath_push_time is not None and now >= _next_filepath_push_time
                run_prune = _next_prune_time is not None and now >= _next_prune_time
                run_verify = _next_verify_time is not None and now >= _next_verify_time

            if run_backup:
                app_logger.info('Scheduler: running scheduled backup')
                _exec_scheduled_backup()
            if run_git:
                app_logger.info('Scheduler: running scheduled git push')
                _exec_scheduled_git_push()
            if run_filepath:
                app_logger.info('Scheduler: running scheduled file path push')
                _exec_scheduled_filepath_push()
            if run_prune:
                app_logger.info('Scheduler: running scheduled prune')
                _exec_scheduled_prune()
            if run_verify:
                app_logger.info('Scheduler: running scheduled verification')
                _exec_scheduled_verify()
        except Exception:
            app_logger.error('Scheduler loop error (will continue):\n%s', traceback.format_exc())

        # Sleep in 5-second chunks so stop events are responsive
        for _ in range(12):
            if _scheduler_stop.is_set():
                break
            _scheduler_stop.wait(5)


def _retry_or_reschedule(task_name, start_func, stop_func, config_enabled_key, config_interval_key):
    """Handle retry backoff on failure or normal reschedule on success."""
    config = db._get_backup_config()
    if not config.get(config_enabled_key):
        stop_func()
        _fail_count[task_name] = 0
        return
    fails = _fail_count[task_name]
    if fails > 0 and fails <= len(_RETRY_DELAYS_MIN):
        retry_minutes = _RETRY_DELAYS_MIN[fails - 1]
        app_logger.warning('Scheduled %s: retry %d/%d in %d minutes',
                           task_name, fails, len(_RETRY_DELAYS_MIN), retry_minutes)
        start_func(retry_minutes / 60.0)
    else:
        # Normal interval (either success or retries exhausted)
        if fails > len(_RETRY_DELAYS_MIN):
            app_logger.error('Scheduled %s: all %d retries exhausted, resuming normal interval',
                             task_name, len(_RETRY_DELAYS_MIN))
            _fail_count[task_name] = 0
        start_func(config[config_interval_key])


def _exec_scheduled_backup():
    """Run backup and reschedule from latest config, with retry on failure."""
    try:
        result = db.backup_database(performed_by='scheduled')
        if result.get('skipped'):
            app_logger.info('Scheduled backup skipped — database unchanged')
        else:
            app_logger.info('Scheduled backup completed: %s (%d bytes, pruned=%d)',
                            result['filename'], result['size'], result['pruned'])
            # Mirror to cloud destinations so they stay in sync with local
            _mirror_local_to_cloud(reason='scheduled backup')
        _fail_count['backup'] = 0
    except Exception as e:
        app_logger.error('Scheduled backup failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        _fail_count['backup'] += 1
    _retry_or_reschedule('backup', _start_backup_timer, _stop_backup_timer,
                         'backup_enabled', 'backup_interval_hours')


def _exec_scheduled_git_push():
    """Run git push and reschedule from latest config, with retry on failure."""
    try:
        result = db.push_backups_to_git()
        if result.get('skipped'):
            app_logger.info('Scheduled git push skipped — backup zip unchanged')
        else:
            app_logger.info('Scheduled git push completed: %d files to %s',
                            result['files_pushed'], result['pushed_to'])
        _fail_count['git_push'] = 0
    except Exception as e:
        app_logger.error('Scheduled git push failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        _fail_count['git_push'] += 1
    _retry_or_reschedule('git_push', _start_git_push_timer, _stop_git_push_timer,
                         'git_enabled', 'git_push_interval_hours')


def _exec_scheduled_filepath_push():
    """Run file path push and reschedule from latest config, with retry on failure."""
    try:
        result = db.push_backups_to_filepath()
        if result.get('skipped'):
            app_logger.info('Scheduled file path push skipped — backup zip unchanged')
        else:
            app_logger.info('Scheduled file path push completed: %d files to %s',
                            result['files_pushed'], result['pushed_to'])
        _fail_count['filepath_push'] = 0
    except Exception as e:
        app_logger.error('Scheduled file path push failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        _fail_count['filepath_push'] += 1
    _retry_or_reschedule('filepath_push', _start_filepath_push_timer, _stop_filepath_push_timer,
                         'filepath_enabled', 'filepath_push_interval_hours')


def _mirror_local_to_cloud(reason='local change'):
    """Push to git and/or file path destinations so their zip mirrors the
    current local backup state. Called after local create/delete/prune
    when the destination has mirroring enabled. Errors are logged but
    do not raise — cloud mirroring is best-effort."""
    try:
        config = db._get_backup_config()
    except Exception:
        return
    if config.get('git_enabled') and config.get('git_repo') and config.get('mirror_to_git'):
        try:
            db.push_backups_to_git()
            app_logger.info('Cloud mirror (git) synced after %s', reason)
        except Exception as e:
            app_logger.warning('Cloud mirror (git) failed after %s: %s', reason, e)
    if config.get('filepath_enabled') and config.get('filepath_path') and config.get('mirror_to_filepath'):
        try:
            db.push_backups_to_filepath()
            app_logger.info('Cloud mirror (filepath) synced after %s', reason)
        except Exception as e:
            app_logger.warning('Cloud mirror (filepath) failed after %s: %s', reason, e)


def _exec_scheduled_prune():
    """Run prune and reschedule from latest config, with retry on failure."""
    try:
        config = db._get_backup_config()
        pruned = db._smart_prune_backups(config['max_backups'])
        if pruned:
            app_logger.info('Scheduled prune completed: removed %d old auto-backups', pruned)
            # Mirror the pruned state to cloud destinations if enabled
            _mirror_local_to_cloud(reason='scheduled prune')
        _fail_count['prune'] = 0
    except Exception as e:
        app_logger.error('Scheduled prune failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        _fail_count['prune'] += 1
    _retry_or_reschedule('prune', _start_prune_timer, _stop_prune_timer,
                         'prune_enabled', 'prune_interval_hours')


def _exec_scheduled_verify():
    """Run backup verification and reschedule."""
    try:
        result = db.verify_backup(rotate=True)
        if result['ok']:
            app_logger.info('Backup verification passed: %s', result['filename'])
        else:
            app_logger.error('BACKUP VERIFICATION FAILED: %s — %s',
                             result['filename'], result['result'])
    except Exception as e:
        app_logger.error('Backup verification error: %s\nTraceback:\n%s', e, traceback.format_exc())
    # Re-arm: verify every 24 hours
    with _scheduler_lock:
        global _next_verify_time
        _next_verify_time = datetime.now() + timedelta(hours=24)


def _start_backup_timer(interval_hours):
    """Schedule the next backup after interval_hours from now."""
    global _next_backup_time
    seconds = max(interval_hours * 3600, 300)  # Minimum 5 minutes
    with _scheduler_lock:
        _next_backup_time = datetime.now() + timedelta(seconds=seconds)
    app_logger.info('Backup scheduler armed: next backup in %s hours', interval_hours)


def _stop_backup_timer():
    """Disable scheduled backups."""
    global _next_backup_time
    with _scheduler_lock:
        _next_backup_time = None


def _start_git_push_timer(interval_hours):
    """Schedule the next git push after interval_hours from now."""
    global _next_git_push_time
    seconds = max(interval_hours * 3600, 300)
    with _scheduler_lock:
        _next_git_push_time = datetime.now() + timedelta(seconds=seconds)
    app_logger.info('Git push scheduler armed: next push in %s hours', interval_hours)


def _stop_git_push_timer():
    """Disable scheduled git pushes."""
    global _next_git_push_time
    with _scheduler_lock:
        _next_git_push_time = None


def _start_filepath_push_timer(interval_hours):
    """Schedule the next file path push after interval_hours from now."""
    global _next_filepath_push_time
    seconds = max(interval_hours * 3600, 300)
    with _scheduler_lock:
        _next_filepath_push_time = datetime.now() + timedelta(seconds=seconds)
    app_logger.info('File path push scheduler armed: next push in %s hours', interval_hours)


def _stop_filepath_push_timer():
    """Disable scheduled file path pushes."""
    global _next_filepath_push_time
    with _scheduler_lock:
        _next_filepath_push_time = None


def _start_prune_timer(interval_hours):
    """Schedule the next prune after interval_hours from now."""
    global _next_prune_time
    seconds = max(interval_hours * 3600, 300)
    with _scheduler_lock:
        _next_prune_time = datetime.now() + timedelta(seconds=seconds)
    app_logger.info('Prune scheduler armed: next prune in %s hours', interval_hours)


def _stop_prune_timer():
    """Disable scheduled prunes."""
    global _next_prune_time
    with _scheduler_lock:
        _next_prune_time = None


def _ensure_scheduler_running():
    """Start the scheduler thread if it isn't already alive (thread-safe)."""
    global _scheduler_thread
    with _scheduler_lock:
        if _scheduler_thread is not None and _scheduler_thread.is_alive():
            return
        _scheduler_stop.clear()
        _scheduler_thread = threading.Thread(target=_scheduler_loop, name='backup-scheduler', daemon=True)
        _scheduler_thread.start()
    app_logger.info('Backup scheduler thread started')


# Restore schedules on startup — if a task is overdue, run it soon instead of
# waiting a full interval (prevents persistent "overdue" after restart).
_startup_config = db._get_backup_config()
if _startup_config.get('backup_enabled'):
    _last_bk = _startup_config.get('last_backup', '')
    _bk_interval = _startup_config['backup_interval_hours']
    if _last_bk:
        _bk_age_hours = (datetime.now() - datetime.strptime(_last_bk, '%Y-%m-%d %H:%M:%S')).total_seconds() / 3600
        if _bk_age_hours > _bk_interval:
            # Only treat as truly overdue if the database has changed since last backup
            _current_hash = db._compute_db_hash()
            _last_hash = _startup_config.get('last_backup_hash', '')
            if _current_hash and _current_hash == _last_hash:
                app_logger.info('Backup age (%.1f hours) exceeds interval but database unchanged — scheduling at normal interval', _bk_age_hours)
                _start_backup_timer(_bk_interval)
            else:
                app_logger.info('Backup overdue on startup (%.1f hours old, database changed), scheduling in 30 seconds', _bk_age_hours)
                _start_backup_timer(30 / 3600)  # ~30 seconds
        else:
            _start_backup_timer(_bk_interval - _bk_age_hours)
    else:
        _start_backup_timer(_bk_interval)
if _startup_config.get('git_enabled') and _startup_config.get('git_repo'):
    _last_gp = _startup_config.get('last_git_push', '')
    _gp_interval = _startup_config['git_push_interval_hours']
    if _last_gp:
        _gp_age_hours = (datetime.now() - datetime.strptime(_last_gp, '%Y-%m-%d %H:%M:%S')).total_seconds() / 3600
        if _gp_age_hours > _gp_interval:
            _start_git_push_timer(30 / 3600)
        else:
            _start_git_push_timer(_gp_interval - _gp_age_hours)
    else:
        _start_git_push_timer(_gp_interval)
if _startup_config.get('filepath_enabled') and _startup_config.get('filepath_path'):
    _last_fp = _startup_config.get('last_filepath_push', '')
    _fp_interval = _startup_config['filepath_push_interval_hours']
    if _last_fp:
        _fp_age_hours = (datetime.now() - datetime.strptime(_last_fp, '%Y-%m-%d %H:%M:%S')).total_seconds() / 3600
        if _fp_age_hours > _fp_interval:
            _start_filepath_push_timer(30 / 3600)
        else:
            _start_filepath_push_timer(_fp_interval - _fp_age_hours)
    else:
        _start_filepath_push_timer(_fp_interval)
if _startup_config.get('prune_enabled'):
    _start_prune_timer(_startup_config['prune_interval_hours'])

# Verification runs every 24 hours regardless of config
_next_verify_time = datetime.now() + timedelta(hours=24)

# Start the single persistent scheduler thread
_ensure_scheduler_running()


@app.route('/backups')
@permission_required('backups')
def backup_list():
    """View backup management page."""
    _ensure_scheduler_running()  # Self-heal if scheduler died
    backups = db.list_backups()
    config = db._get_backup_config()
    with _scheduler_lock:
        next_backup = _next_backup_time.strftime('%Y-%m-%d %H:%M:%S') if _next_backup_time else None
        next_push = _next_git_push_time.strftime('%Y-%m-%d %H:%M:%S') if _next_git_push_time else None
        next_filepath = _next_filepath_push_time.strftime('%Y-%m-%d %H:%M:%S') if _next_filepath_push_time else None
        next_prune = _next_prune_time.strftime('%Y-%m-%d %H:%M:%S') if _next_prune_time else None
    scheduler_alive = _scheduler_thread is not None and _scheduler_thread.is_alive()
    health = db.get_backup_health()
    # Show verification failure as an error flash (only when failed)
    if config.get('last_verify_time') and not config.get('last_verify_ok'):
        verify_msg = f'Backup verification FAILED: {config.get("last_verify_file", "unknown")}'
        if config.get('last_verify_result'):
            verify_msg += f' — {config["last_verify_result"]}'
        flash(verify_msg, 'error')
    cloud_files = set(config.get('last_cloud_backup_files', []))
    filepath_files = set(config.get('last_filepath_backup_files', []))
    return render_template('backups.html', backups=backups, config=config,
                           next_backup_time=next_backup, next_git_push_time=next_push,
                           next_filepath_push_time=next_filepath,
                           next_prune_time=next_prune, scheduler_alive=scheduler_alive,
                           backup_health=health, cloud_backup_files=cloud_files,
                           filepath_backup_files=filepath_files)


@app.route('/backups/create', methods=['POST'])
@permission_required('backups')
def backup_create():
    """Trigger a manual database backup."""
    try:
        result = db.backup_database(performed_by=current_username(), manual=True)
        app_logger.info('Manual backup created: %s (%d bytes, pruned=%d) by=%s',
                        result['filename'], result['size'], result['pruned'],
                        current_username())
        flash(f'Backup created: {result["filename"]}', 'success')
        # Mirror to cloud destinations if enabled
        _mirror_local_to_cloud(reason='manual backup')
    except Exception as e:
        app_logger.error('Manual backup failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Backup failed: {e}', 'error')
    return redirect(url_for('backup_list'))


@app.route('/backups/upload', methods=['POST'])
@permission_required('backups')
def backup_upload():
    """Restore database from an uploaded .db or .zip backup file."""
    MAX_UPLOAD_MB = 500
    file = request.files.get('backup_file')
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('backup_list'))
    is_zip = file.filename.endswith('.zip')
    is_db = file.filename.endswith('.db')
    if not is_zip and not is_db:
        flash('Invalid file type. Please upload a .db or .zip backup file.', 'error')
        return redirect(url_for('backup_list'))
    # Validate magic bytes before saving to disk
    header = file.read(16)
    file.seek(0)
    if is_db and header[:16] != b'SQLite format 3\x00':
        flash('Invalid file: not a valid SQLite database.', 'error')
        return redirect(url_for('backup_list'))
    if is_zip and header[:4] != b'PK\x03\x04':
        flash('Invalid file: not a valid ZIP archive.', 'error')
        return redirect(url_for('backup_list'))
    # Check file size (read content length or measure stream)
    file.seek(0, 2)  # seek to end
    file_size = file.tell()
    file.seek(0)
    if file_size > MAX_UPLOAD_MB * 1024 * 1024:
        flash(f'File too large ({file_size // (1024*1024)} MB). Maximum is {MAX_UPLOAD_MB} MB.', 'error')
        return redirect(url_for('backup_list'))
    try:
        # Save uploaded file to backup dir
        backup_dir = db._get_backup_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ext = '.zip' if is_zip else '.db'
        dest_filename = f'manual_backup_{timestamp}_uploaded{ext}'
        dest_path = os.path.join(backup_dir, dest_filename)
        file.save(dest_path)

        # For .db files, run compatibility check before restore
        if is_db:
            compat = db.validate_backup_compatibility(dest_path)
            if not compat['compatible']:
                error_detail = '; '.join(compat['errors'])
                flash(f'Backup is not compatible: {error_detail}', 'error')
                try:
                    os.remove(dest_path)
                except OSError:
                    pass
                return redirect(url_for('backup_list'))

        # Restore from the uploaded file
        result = db.restore_database(dest_filename)
        app_logger.info('Database restored from upload: %s (safety: %s) by=%s',
                        dest_filename, result['safety_backup'], current_username())
        msg = f'Database restored from uploaded file. Safety backup: {result["safety_backup"]}'
        if result.get('warnings'):
            msg += f' ({len(result["warnings"])} compatibility warning{"s" if len(result["warnings"]) != 1 else ""})'
        flash(msg, 'success')
        for w in result.get('warnings', []):
            flash(w, 'warning')
    except ValueError as e:
        app_logger.error('Upload restore failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Restore failed: {e}', 'error')
    except Exception as e:
        app_logger.error('Upload restore failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Restore failed: {e}', 'error')
    return redirect(url_for('backup_list'))



@app.route('/backups/config', methods=['POST'])
@permission_required('backups')
def backup_config():
    """Update all backup configuration settings."""
    config = db._get_backup_config()

    # Local backup settings
    backup_dir = request.form.get('backup_dir', '').strip()
    if backup_dir:
        if not os.path.isabs(backup_dir):
            flash('Backup directory must be an absolute path.', 'error')
            return redirect(url_for('backup_list'))
        os.makedirs(backup_dir, exist_ok=True)
        if not os.access(backup_dir, os.W_OK):
            flash(f'Backup directory is not writable: {backup_dir}', 'error')
            return redirect(url_for('backup_list'))
        config['backup_dir'] = backup_dir

    try:
        config['max_backups'] = max(1, int(request.form.get('max_backups', 5)))
    except (ValueError, TypeError):
        config['max_backups'] = 5

    config['backup_enabled'] = '1' in request.form.getlist('backup_enabled')
    config['include_uploads'] = '1' in request.form.getlist('include_uploads')
    try:
        config['backup_interval_hours'] = max(0.1, float(request.form.get('backup_interval_hours', 24)))
    except (ValueError, TypeError):
        config['backup_interval_hours'] = 24

    # Prune settings
    config['prune_enabled'] = '1' in request.form.getlist('prune_enabled')
    try:
        config['prune_interval_hours'] = max(0.1, float(request.form.get('prune_interval_hours', 24)))
    except (ValueError, TypeError):
        config['prune_interval_hours'] = 24

    # Mirror toggles — sync cloud zip to match local backup state
    config['mirror_to_git'] = '1' in request.form.getlist('mirror_to_git')
    config['mirror_to_filepath'] = '1' in request.form.getlist('mirror_to_filepath')

    # Git push settings
    config['git_enabled'] = '1' in request.form.getlist('git_enabled')
    git_repo = request.form.get('git_repo', '').strip()
    if git_repo:
        # Basic sanity check: must look like a git URL or path
        valid_repo = (git_repo.startswith(('https://', 'http://', 'git@', 'ssh://'))
                      or git_repo.endswith('.git')
                      or '/' in git_repo)
        if not valid_repo:
            flash('Git repository must be a valid URL (https://...) or SSH path (git@...).', 'error')
            return redirect(url_for('backup_list'))
    config['git_repo'] = git_repo
    config['git_branch'] = request.form.get('git_branch', 'backups').strip() or 'backups'
    config['git_token'] = request.form.get('git_token', '').strip()
    config['git_encryption_password'] = request.form.get('git_encryption_password', '').strip()
    try:
        config['git_push_interval_hours'] = max(0.1, float(request.form.get('git_push_interval_hours', 24)))
    except (ValueError, TypeError):
        config['git_push_interval_hours'] = 24

    # File path backup settings
    config['filepath_enabled'] = '1' in request.form.getlist('filepath_enabled')
    filepath_path = request.form.get('filepath_path', '').strip()
    if filepath_path:
        if not os.path.isabs(filepath_path) and not filepath_path.startswith('\\\\'):
            flash('File path must be an absolute path (e.g. /mnt/backup or \\\\server\\share).', 'error')
            return redirect(url_for('backup_list'))
        # Try to create and validate writability
        try:
            os.makedirs(filepath_path, exist_ok=True)
        except OSError as e:
            flash(f'Cannot access file path: {e}', 'error')
            return redirect(url_for('backup_list'))
        if not os.access(filepath_path, os.W_OK):
            flash(f'File path is not writable: {filepath_path}', 'error')
            return redirect(url_for('backup_list'))
        config['filepath_path'] = filepath_path
    elif not filepath_path and config.get('filepath_path'):
        # Cleared the path — disable if it was set
        config['filepath_path'] = ''
    config['filepath_encryption_password'] = request.form.get('filepath_encryption_password', '').strip()
    try:
        config['filepath_push_interval_hours'] = max(0.1, float(request.form.get('filepath_push_interval_hours', 24)))
    except (ValueError, TypeError):
        config['filepath_push_interval_hours'] = 24

    db.save_backup_config(config)

    # Manage backup timer
    if config['backup_enabled']:
        _start_backup_timer(config['backup_interval_hours'])
        app_logger.info('Backup schedule enabled: every %s hours by=%s',
                        config['backup_interval_hours'], current_username())
    else:
        _stop_backup_timer()

    # Manage git push timer
    if config['git_enabled'] and config['git_repo']:
        _start_git_push_timer(config['git_push_interval_hours'])
        app_logger.info('Git push schedule enabled: every %s hours to %s by=%s',
                        config['git_push_interval_hours'], config['git_repo'], current_username())
    else:
        _stop_git_push_timer()

    # Manage file path push timer
    if config['filepath_enabled'] and config.get('filepath_path'):
        _start_filepath_push_timer(config['filepath_push_interval_hours'])
        app_logger.info('File path push schedule enabled: every %s hours to %s by=%s',
                        config['filepath_push_interval_hours'], config['filepath_path'], current_username())
    else:
        _stop_filepath_push_timer()

    # Manage prune timer
    if config['prune_enabled']:
        _start_prune_timer(config['prune_interval_hours'])
        app_logger.info('Prune schedule enabled: every %s hours by=%s',
                        config['prune_interval_hours'], current_username())
    else:
        _stop_prune_timer()

    # Ensure the scheduler thread is alive (recovers if it died)
    _ensure_scheduler_running()

    app_logger.info('Backup config updated by=%s', current_username())
    flash('Backup configuration saved.', 'success')
    return redirect(url_for('backup_list'))


@app.route('/backups/export-encryption-key')
@permission_required('backups')
def backup_export_encryption_key():
    """Download the cloud backup encryption key as a text file."""
    config = db._get_backup_config()
    key = config.get('git_encryption_password', '').strip()
    if not key:
        flash('No encryption password is configured.', 'error')
        return redirect(url_for('backup_list'))

    app_logger.info('Encryption key exported by=%s', current_username())
    content = (
        'HP Connectivity Inventory — Backup Encryption Key\n'
        '==================================================\n\n'
        f'Encryption Password: {key}\n\n'
        'Store this file in a safe location. You will need this\n'
        'password to decrypt cloud backup files (.zip) if the\n'
        'database or configuration is lost.\n\n'
        'To decrypt a backup manually:\n'
        '  Use any tool that supports AES-256 encrypted ZIP files\n'
        '  (e.g. 7-Zip, WinZip, or Python pyzipper library).\n'
    )
    response = app.response_class(content, mimetype='text/plain')
    response.headers['Content-Disposition'] = 'attachment; filename="backup_encryption_key.txt"'
    return response


@app.route('/backups/config/reset', methods=['POST'])
@permission_required('backups')
def backup_config_reset():
    """Reset backup configuration to factory defaults (preserves git credentials)."""
    current = db._get_backup_config()
    defaults = db.get_default_backup_config()
    # Preserve git credentials and repo settings — user shouldn't have to re-enter these
    defaults['git_repo'] = current.get('git_repo', '')
    defaults['git_branch'] = current.get('git_branch', 'backups')
    defaults['git_token'] = current.get('git_token', '')
    defaults['filepath_path'] = current.get('filepath_path', '')
    # Preserve timestamps
    defaults['last_backup'] = current.get('last_backup', '')
    defaults['last_git_push'] = current.get('last_git_push', '')
    defaults['last_filepath_push'] = current.get('last_filepath_push', '')
    defaults['last_backup_hash'] = current.get('last_backup_hash', '')
    db.save_backup_config(defaults)
    _stop_backup_timer()
    _stop_git_push_timer()
    _stop_filepath_push_timer()
    _stop_prune_timer()
    app_logger.info('Backup config reset to defaults by=%s', current_username())
    flash('Backup configuration reset to defaults.', 'success')
    return redirect(url_for('backup_list'))


@app.route('/backups/browse-directory')
@permission_required('backups')
def backup_browse_directory():
    """API: list sub-directories at a given path for the folder picker."""
    path = request.args.get('path', '').strip()

    # Default starting points by platform
    if not path:
        if sys.platform == 'win32':
            # List drive letters on Windows
            import string
            drives = []
            for letter in string.ascii_uppercase:
                drive = f'{letter}:\\'
                if os.path.isdir(drive):
                    drives.append({'name': drive, 'path': drive})
            return jsonify({'ok': True, 'path': '', 'parent': '', 'entries': drives})
        else:
            path = '/'

    # Normalize and resolve
    path = os.path.normpath(path)
    if not os.path.isabs(path):
        return jsonify({'ok': False, 'error': 'Path must be absolute.'})
    if not os.path.isdir(path):
        return jsonify({'ok': False, 'error': f'Directory not found: {path}'})

    try:
        entries = []
        for name in sorted(os.listdir(path)):
            full = os.path.join(path, name)
            if os.path.isdir(full) and not name.startswith('.'):
                entries.append({'name': name, 'path': full})
        parent = os.path.dirname(path) if path != os.path.dirname(path) else ''
        return jsonify({'ok': True, 'path': path, 'parent': parent, 'entries': entries})
    except PermissionError:
        return jsonify({'ok': False, 'error': f'Permission denied: {path}'})
    except OSError as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/backups/push', methods=['POST'])
@permission_required('backups')
def backup_push_git():
    """Manually trigger a git push of backup bundle."""
    try:
        result = db.push_backups_to_git()
        app_logger.info('Manual git push: %d files to %s by=%s',
                        result['files_pushed'], result['pushed_to'], current_username())
        flash(f'Backups pushed to git: {result["files_pushed"]} .db files pushed to {result["pushed_to"]}', 'success')
    except Exception as e:
        app_logger.error('Git push failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Git push failed: {e}', 'error')
    return redirect(url_for('backup_list'))


@app.route('/backups/local/list')
@permission_required('backups')
def backup_local_list():
    """API: list .db files in the local backup directory."""
    try:
        backups = db.list_backups()
        return jsonify({'ok': True, 'backups': backups})
    except Exception as e:
        app_logger.error('Local backup list failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/backups/git/list')
@permission_required('backups')
def backup_git_list():
    """API: list .db files available in the git backup zip."""
    try:
        entries = db.list_git_backups()
        return jsonify({'ok': True, 'backups': entries})
    except Exception as e:
        app_logger.error('Git backup list failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/backups/git/restore', methods=['POST'])
@permission_required('backups')
def backup_git_restore():
    """Restore database from a file in the git backup zip. Requires admin password."""
    filename = request.form.get('filename', '').strip()
    if not filename:
        flash('No file selected.', 'error')
        return redirect(url_for('backup_list'))

    # Require admin password re-entry for cloud restore
    admin_password = request.form.get('admin_password', '').strip()
    if not admin_password:
        flash('Admin password is required to restore from cloud backup.', 'error')
        return redirect(url_for('backup_list'))
    user = db.authenticate_user(current_username(), admin_password)
    if not user or user.get('role') != 'admin':
        app_logger.warning('Cloud restore blocked: invalid admin password by=%s', current_username())
        flash('Invalid admin password. Cloud restore requires admin authentication.', 'error')
        return redirect(url_for('backup_list'))

    try:
        result = db.restore_from_git(filename)
        app_logger.info('Database restored from git: %s (safety: %s) by=%s',
                        result['restored_from'], result['safety_backup'], current_username())
        flash(f'Database restored from git backup: {filename}. Safety backup: {result["safety_backup"]}', 'success')
    except Exception as e:
        app_logger.error('Git restore failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Restore from git failed: {e}', 'error')
    return redirect(url_for('backup_list'))


@app.route('/backups/filepath/push', methods=['POST'])
@permission_required('backups')
def backup_push_filepath():
    """Manually trigger a file path push of backup bundle."""
    try:
        result = db.push_backups_to_filepath()
        app_logger.info('Manual file path push: %d files to %s by=%s',
                        result['files_pushed'], result['pushed_to'], current_username())
        flash(f'Backups pushed to file path: {result["files_pushed"]} .db files copied to {result["pushed_to"]}', 'success')
    except Exception as e:
        app_logger.error('File path push failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'File path push failed: {e}', 'error')
    return redirect(url_for('backup_list'))


@app.route('/backups/filepath/list')
@permission_required('backups')
def backup_filepath_list():
    """API: list .db files available in the file path backup zip."""
    try:
        entries = db.list_filepath_backups()
        return jsonify({'ok': True, 'backups': entries})
    except Exception as e:
        app_logger.error('File path backup list failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/backups/filepath/restore', methods=['POST'])
@permission_required('backups')
def backup_filepath_restore():
    """Restore database from a file in the file path backup zip. Requires admin password."""
    filename = request.form.get('filename', '').strip()
    if not filename:
        flash('No file selected.', 'error')
        return redirect(url_for('backup_list'))

    admin_password = request.form.get('admin_password', '').strip()
    if not admin_password:
        flash('Admin password is required to restore from file path backup.', 'error')
        return redirect(url_for('backup_list'))
    user = db.authenticate_user(current_username(), admin_password)
    if not user or user.get('role') != 'admin':
        app_logger.warning('File path restore blocked: invalid admin password by=%s', current_username())
        flash('Invalid admin password. Restore requires admin authentication.', 'error')
        return redirect(url_for('backup_list'))

    try:
        result = db.restore_from_filepath(filename)
        app_logger.info('Database restored from file path: %s (safety: %s) by=%s',
                        result['restored_from'], result['safety_backup'], current_username())
        flash(f'Database restored from file path backup: {filename}. Safety backup: {result["safety_backup"]}', 'success')
    except Exception as e:
        app_logger.error('File path restore failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Restore from file path failed: {e}', 'error')
    return redirect(url_for('backup_list'))


@app.route('/backups/<filename>/delete', methods=['POST'])
@permission_required('backups')
def backup_delete(filename):
    """Delete a backup file."""
    try:
        db.delete_backup(filename)
        app_logger.info('Backup deleted: %s by=%s', filename, current_username())
        flash(f'Backup deleted: {filename}', 'success')
        # Mirror the removal to cloud destinations if enabled
        _mirror_local_to_cloud(reason='manual delete')
    except Exception as e:
        app_logger.error('Backup delete failed: file=%s error=%s\nTraceback:\n%s', filename, e, traceback.format_exc())
        flash(f'Error deleting backup: {e}', 'error')
    return redirect(url_for('backup_list'))


@app.route('/backups/<filename>/download')
@permission_required('backups')
def backup_download(filename):
    """Download a backup file."""
    if not db._is_backup_file(filename) or '..' in filename:
        flash('Invalid backup file.', 'error')
        return redirect(url_for('backup_list'))
    backup_dir = db._get_backup_dir()
    path = os.path.join(backup_dir, filename)
    if not os.path.isfile(path):
        flash('Backup file not found.', 'error')
        return redirect(url_for('backup_list'))
    app_logger.info('Backup downloaded: %s by=%s', filename, current_username())
    return send_file(path, as_attachment=True, download_name=filename)


@app.route('/backups/<filename>/restore', methods=['POST'])
@permission_required('backups')
def backup_restore(filename):
    """Restore the database from a backup file."""
    try:
        result = db.restore_database(filename)
        app_logger.info('Database restored from %s (safety backup: %s) by=%s',
                        result['restored_from'], result['safety_backup'], current_username())
        msg = f'Database restored from {filename}. A safety backup was created: {result["safety_backup"]}'
        if result.get('warnings'):
            msg += f' ({len(result["warnings"])} compatibility warning{"s" if len(result["warnings"]) != 1 else ""})'
        flash(msg, 'success')
        for w in result.get('warnings', []):
            flash(w, 'warning')
    except FileNotFoundError:
        flash('Backup file not found.', 'error')
    except ValueError as e:
        app_logger.error('Restore failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Restore failed: {e}', 'error')
    except Exception as e:
        app_logger.error('Restore failed: %s by=%s\nTraceback:\n%s', e, current_username(), traceback.format_exc())
        flash(f'Restore failed: {e}', 'error')
    return redirect(url_for('backup_list'))

# ---------------------------------------------------------------------------
# Product Reference (printer/device spec catalog)
# ---------------------------------------------------------------------------

@app.route('/reference')
def product_reference_list():
    search = request.args.get('q', '')
    refs = db.get_all_product_references(search)
    inv_counts = db.get_inventory_counts_by_codename()
    return render_template('product_reference.html', refs=refs, search=search, inv_counts=inv_counts)


@app.route('/reference/add', methods=['GET', 'POST'])
@permission_required('references')
def product_reference_add():
    if request.method == 'POST':
        codename = request.form.get('codename', '').strip()
        if not codename:
            flash('Codename is required.', 'error')
            return render_template('product_reference_form.html', ref=None, current_year=str(datetime.now().year))

        db.add_product_reference(
            codename=codename,
            model_name=request.form.get('model_name', '').strip(),
            wifi_gen=request.form.get('wifi_gen', '').strip(),
            year=request.form.get('year', '').strip(),
            chip_manufacturer=request.form.get('chip_manufacturer', '').strip(),
            chip_codename=request.form.get('chip_codename', '').strip(),
            fw_codebase=request.form.get('fw_codebase', '').strip(),
            print_technology=request.form.get('print_technology', '').strip(),
            cartridge_toner=request.form.get('cartridge_toner', '').strip(),
            predecessor=request.form.get('predecessor', '').strip(),
        )
        app_logger.info('Product reference added: codename="%s" by=%s', codename, current_username())
        flash(f'Product reference "{codename}" added.', 'success')
        return redirect(url_for('product_reference_list'))

    return render_template('product_reference_form.html', ref=None, current_year=str(datetime.now().year))


@app.route('/reference/<int:ref_id>/edit', methods=['GET', 'POST'])
@permission_required('references')
def product_reference_edit(ref_id):
    ref = db.get_product_reference(ref_id)
    if not ref:
        flash('Product reference not found.', 'error')
        return redirect(url_for('product_reference_list'))

    if request.method == 'POST':
        codename = request.form.get('codename', '').strip()
        if not codename:
            flash('Codename is required.', 'error')
            return render_template('product_reference_form.html', ref=ref)

        db.update_product_reference(
            ref_id=ref_id,
            codename=codename,
            model_name=request.form.get('model_name', '').strip(),
            wifi_gen=request.form.get('wifi_gen', '').strip(),
            year=request.form.get('year', '').strip(),
            chip_manufacturer=request.form.get('chip_manufacturer', '').strip(),
            chip_codename=request.form.get('chip_codename', '').strip(),
            fw_codebase=request.form.get('fw_codebase', '').strip(),
            print_technology=request.form.get('print_technology', '').strip(),
            cartridge_toner=request.form.get('cartridge_toner', '').strip(),
            predecessor=request.form.get('predecessor', '').strip(),
        )
        app_logger.info('Product reference updated: ref_id=%d codename="%s" by=%s', ref_id, codename, current_username())
        flash(f'Product reference "{codename}" updated.', 'success')
        return redirect(url_for('product_reference_list'))

    return render_template('product_reference_form.html', ref=ref)


@app.route('/api/reference/<int:ref_id>', methods=['PATCH'])
def api_reference_update(ref_id):
    """Inline edit API — update a single field on a product reference."""
    if not has_permission('references'):
        return jsonify({'error': 'Permission denied'}), 403
    ref = db.get_product_reference(ref_id)
    if not ref:
        return jsonify({'error': 'Not found'}), 404
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    allowed = {'codename', 'model_name', 'wifi_gen', 'year', 'chip_manufacturer',
               'chip_codename', 'fw_codebase', 'print_technology', 'cartridge_toner', 'predecessor'}
    updates = {k: v.strip() for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({'error': 'No valid fields'}), 400
    # Merge with existing values
    merged = {k: ref[k] for k in allowed}
    merged.update(updates)
    if not merged.get('codename'):
        return jsonify({'error': 'Codename is required'}), 400
    db.update_product_reference(ref_id=ref_id, **merged)
    app_logger.info('Product reference inline edit: ref_id=%d fields=%s by=%s', ref_id, list(updates.keys()), current_username())
    return jsonify({'ok': True})


@app.route('/reference/<int:ref_id>/delete', methods=['POST'])
@permission_required('references')
def product_reference_delete(ref_id):
    db.delete_product_reference(ref_id)
    app_logger.info('Product reference deleted: ref_id=%d by=%s', ref_id, current_username())
    flash('Product reference deleted.', 'success')
    return redirect(url_for('product_reference_list'))


HEADER_MAP = {
    # Codename (required field)
    'codename': 'codename',
    'code name': 'codename',
    'product codename': 'codename',
    'product': 'codename',
    # Model name
    'model name': 'model_name',
    'model_name': 'model_name',
    'model': 'model_name',
    # Wi-Fi generation
    'wi-fi gen': 'wifi_gen',
    'wifi gen': 'wifi_gen',
    'wifi_gen': 'wifi_gen',
    'wi-fi generation': 'wifi_gen',
    'wifi generation': 'wifi_gen',
    'wireless gen': 'wifi_gen',
    # Year
    'year': 'year',
    'release year': 'year',
    # Chip manufacturer
    'wireless chip set manufacturer': 'chip_manufacturer',
    'wireless chipset manufacturer': 'chip_manufacturer',
    'chip manufacturer': 'chip_manufacturer',
    'chip_manufacturer': 'chip_manufacturer',
    'chip vendor': 'chip_manufacturer',
    'wireless chip vendor': 'chip_manufacturer',
    # Chip codename
    'wireless chipset codename': 'chip_codename',
    'chip codename': 'chip_codename',
    'chip_codename': 'chip_codename',
    # Firmware codebase
    'fw codebase': 'fw_codebase',
    'fw_codebase': 'fw_codebase',
    'firmware codebase': 'fw_codebase',
    'codebase': 'fw_codebase',
    # Print technology
    'print technology': 'print_technology',
    'print_technology': 'print_technology',
    'technology': 'print_technology',
    # Cartridge/Toner
    'cartridge/toner': 'cartridge_toner',
    'cartridge_toner': 'cartridge_toner',
    'cartridge': 'cartridge_toner',
    'toner': 'cartridge_toner',
    'cartridge / toner': 'cartridge_toner',
    # Predecessor
    'predecessor': 'predecessor',
    'predecessor codename': 'predecessor',
    'predecessor_codename': 'predecessor',
    'previous codename': 'predecessor',
}


def _import_seed_data():
    """Import seed CSV (upsert) and attach seed images to wiki pages."""
    import csv as _csv
    seed_dir = os.path.join(BUNDLE_DIR, 'seed_data')
    csv_path = os.path.join(seed_dir, 'product_reference.csv')

    if not os.path.isfile(csv_path):
        flash('Seed data not found. No seed_data/product_reference.csv in the application bundle.', 'error')
        return redirect(url_for('product_reference_list'))

    added = 0
    updated = 0
    skipped = 0
    try:
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = _csv.DictReader(f)
            if reader.fieldnames is None:
                flash('Seed CSV has no headers.', 'error')
                return redirect(url_for('product_reference_list'))
            for row in reader:
                norm = {k.strip().lower(): v.strip() for k, v in row.items() if k}
                codename = norm.get('codename', '').strip()
                if not codename:
                    skipped += 1
                    continue
                _ref_id, action = db.upsert_product_reference(
                    codename=codename,
                    model_name=norm.get('model name', norm.get('model_name', '')),
                    wifi_gen=norm.get('wi-fi gen', norm.get('wifi gen', norm.get('wifi_gen', ''))),
                    year=norm.get('year', ''),
                    chip_manufacturer=norm.get('wireless chip set manufacturer',
                                     norm.get('chip manufacturer', norm.get('chip_manufacturer', ''))),
                    chip_codename=norm.get('wireless chipset codename',
                                  norm.get('chip codename', norm.get('chip_codename', ''))),
                    fw_codebase=norm.get('fw codebase', norm.get('fw_codebase', '')),
                    print_technology=norm.get('print technology', norm.get('print_technology', '')),
                    cartridge_toner=norm.get('cartridge/toner', norm.get('cartridge_toner', '')),
                )
                if action == 'added':
                    added += 1
                else:
                    updated += 1
    except Exception as e:
        app_logger.error('Seed CSV import failed: %s\n%s', e, traceback.format_exc())
        flash(f'Seed import failed: {e}', 'error')
        return redirect(url_for('product_reference_list'))

    # Phase 2: attach seed images to wiki pages
    images_attached = 0
    zip_path = os.path.join(seed_dir, 'printer_images.zip')
    if os.path.isfile(zip_path):
        try:
            from database import _seed_wiki_images
            images_attached = _seed_wiki_images(zip_path)
        except Exception as e:
            app_logger.error('Seed image attachment failed: %s\n%s', e, traceback.format_exc())
            flash(f'Image attachment partially failed: {e}', 'warning')

    parts = []
    if added:
        parts.append(f'{added} added')
    if updated:
        parts.append(f'{updated} updated')
    if skipped:
        parts.append(f'{skipped} skipped')
    msg = f'Seed import: {", ".join(parts)}.'
    if images_attached:
        msg += f' {images_attached} images attached to wiki pages.'
    app_logger.info('Seed import completed: added=%d updated=%d skipped=%d images=%d by=%s',
                    added, updated, skipped, images_attached, current_username())
    flash(msg, 'success')
    return redirect(url_for('product_reference_list'))


@app.route('/reference/seed', methods=['POST'])
@permission_required('references')
def product_reference_seed():
    """Import seed data from the application bundle."""
    return _import_seed_data()


@app.route('/reference/import', methods=['POST'])
@permission_required('references')
def product_reference_import():
    """Import product references from an uploaded .xlsx, .csv, or .zip file.

    ZIP files are expected to contain a product_reference.csv at the root
    and an optional images/ directory with sub-folders named by codename,
    each containing attachment files. This matches the ZIP export format.
    """
    file = request.files.get('import_file')
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('product_reference_list'))

    filename = file.filename.lower()
    if not filename.endswith(('.xlsx', '.csv', '.zip')):
        flash('Unsupported file type. Use .xlsx, .csv, or .zip', 'error')
        return redirect(url_for('product_reference_list'))

    try:
        import_mode = request.form.get('import_mode', 'add')
        if import_mode == 'overwrite':
            db.clear_all_product_references()

        imported = 0
        skipped = 0

        def _map_headers(raw):
            """Map raw header names to DB columns, warn about unrecognized ones."""
            mapped = [HEADER_MAP.get(str(h).strip().lower()) for h in raw]
            if not any(m == 'codename' for m in mapped):
                flash('Warning: No "Codename" column found. All rows will be skipped.', 'warning')
            unrecognized = [str(h).strip() for h, m in zip(raw, mapped)
                            if m is None and str(h).strip()]
            if unrecognized:
                flash(f'Unrecognized columns ignored: {", ".join(unrecognized)}', 'warning')
            return mapped

        def _import_csv_text(text):
            """Import from CSV text, return (imported, skipped) counts."""
            nonlocal imported, skipped
            delimiter = '\t' if '\t' in text[:2048] else ','
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            raw_headers = next(reader)
            headers = _map_headers(raw_headers)
            for row in reader:
                if not any(cell.strip() for cell in row):
                    continue
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = val.strip()
                codename = record.get('codename', '').strip()
                if not codename:
                    skipped += 1
                    continue
                db.add_product_reference(**record)
                imported += 1

        if filename.endswith('.zip'):
            # --- ZIP bundle import (CSV + printer_images.zip) ---
            import tempfile
            raw_data = file.read()
            zf = zipfile.ZipFile(io.BytesIO(raw_data))

            # Find and import the CSV
            csv_names = [n for n in zf.namelist() if n.lower().endswith('.csv')]
            if not csv_names:
                flash('ZIP file does not contain a .csv file.', 'error')
                return redirect(url_for('product_reference_list'))
            csv_text = zf.read(csv_names[0]).decode('utf-8-sig')
            _import_csv_text(csv_text)

            # Extract printer_images.zip (if present) and attach via _seed_wiki_images
            images_attached = 0
            zip_names = [n for n in zf.namelist() if n.lower().endswith('.zip')]
            for inner_name in zip_names:
                inner_data = zf.read(inner_name)
                with tempfile.NamedTemporaryFile(suffix='.zip', delete=False) as tmp:
                    tmp.write(inner_data)
                    tmp_path = tmp.name
                try:
                    from database import _seed_wiki_images
                    images_attached += _seed_wiki_images(tmp_path)
                finally:
                    os.unlink(tmp_path)

            zf.close()
            msg = f'Imported {imported} product{"s" if imported != 1 else ""}.'
            if skipped:
                msg += f' {skipped} rows skipped (no codename).'
            if images_attached:
                msg += f' {images_attached} images attached.'
            flash(msg, 'success')

        elif filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows()
            raw_headers = [cell.value or '' for cell in next(rows)]
            headers = _map_headers(raw_headers)

            for row in rows:
                values = [cell.value for cell in row]
                if not any(v is not None and str(v).strip() for v in values):
                    continue
                record = {}
                for i, val in enumerate(values):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = str(val).strip() if val is not None else ''
                codename = record.get('codename', '').strip()
                if not codename:
                    skipped += 1
                    continue
                db.add_product_reference(**record)
                imported += 1
            wb.close()
            flash(f'Imported {imported} product{"s" if imported != 1 else ""}.'
                  + (f' {skipped} rows skipped (no codename).' if skipped else ''), 'success')
        else:
            raw = file.read()
            text = raw.decode('utf-8-sig')
            _import_csv_text(text)
            flash(f'Imported {imported} product{"s" if imported != 1 else ""}.'
                  + (f' {skipped} rows skipped (no codename).' if skipped else ''), 'success')

        app_logger.info('Product reference import: %d imported, %d skipped, file="%s" by=%s',
                        imported, skipped, file.filename, current_username())
    except Exception as e:
        app_logger.error('Product reference import failed: %s\nTraceback:\n%s', e, traceback.format_exc())
        flash(f'Import failed: {e}', 'error')

    return redirect(url_for('product_reference_list'))


@app.route('/reference/export')
@permission_required('references')
def product_reference_export():
    """Export all product references as a .csv download."""
    import csv, io
    refs = db.get_all_product_references()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Codename', 'Model Name', 'Print Technology', 'Cartridge/Toner', 'Wi-Fi Gen', 'Year',
                     'Wireless Chip Set Manufacturer', 'Wireless Chipset Codename', 'FW Codebase',
                     'Predecessor'])
    for r in refs:
        writer.writerow([r['codename'], r['model_name'], r['print_technology'],
                         r.get('cartridge_toner', ''), r['wifi_gen'], r['year'],
                         r['chip_manufacturer'], r['chip_codename'], r['fw_codebase'],
                         r.get('predecessor', '')])
    csv_bytes = output.getvalue().encode('utf-8-sig')
    app_logger.info('Product reference CSV export: %d refs by=%s', len(refs), current_username())
    return Response(csv_bytes, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=product_reference.csv'})


@app.route('/reference/export/xlsx')
@permission_required('references')
def product_reference_export_xlsx():
    """Export all product references as an .xlsx download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl is required for Excel export. Install with: pip install openpyxl', 'error')
        return redirect(url_for('product_reference_list'))

    refs = db.get_all_product_references()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product Reference'

    headers = ['Codename', 'Model Name', 'Print Technology', 'Cartridge/Toner', 'Wi-Fi Gen', 'Year',
               'Wireless Chip Set Manufacturer', 'Wireless Chipset Codename', 'FW Codebase',
               'Predecessor']
    ws.append(headers)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for r in refs:
        ws.append([r['codename'], r['model_name'], r['print_technology'],
                   r.get('cartridge_toner', ''), r['wifi_gen'], r['year'],
                   r['chip_manufacturer'], r['chip_codename'], r['fw_codebase'],
                   r.get('predecessor', '')])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=product_reference.xlsx'}
    )


@app.route('/reference/export/zip')
@permission_required('references')
def product_reference_export_zip():
    """Export product references as a ZIP bundle with CSV + printer_images.zip.

    Mirrors the seed_data structure so the ZIP can be imported on another
    machine using the same fuzzy-matching image attachment logic.
    """
    refs = db.get_all_product_references()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as outer:
        # --- product_reference.csv ---
        csv_buf = io.StringIO()
        writer = csv.writer(csv_buf)
        writer.writerow(['Codename', 'Model Name', 'Print Technology', 'Cartridge/Toner',
                         'Wi-Fi Gen', 'Year', 'Wireless Chip Set Manufacturer',
                         'Wireless Chipset Codename', 'FW Codebase', 'Predecessor'])
        for r in refs:
            writer.writerow([r['codename'], r['model_name'], r['print_technology'],
                             r.get('cartridge_toner', ''), r['wifi_gen'], r['year'],
                             r['chip_manufacturer'], r['chip_codename'], r['fw_codebase'],
                             r.get('predecessor', '')])
        outer.writestr('product_reference.csv', csv_buf.getvalue().encode('utf-8-sig'))

        # --- printer_images.zip (inner ZIP with wiki attachment images) ---
        img_buf = io.BytesIO()
        img_count = 0
        with zipfile.ZipFile(img_buf, 'w', zipfile.ZIP_DEFLATED) as inner:
            for ref in refs:
                attachments = db.get_wiki_attachments(ref['ref_id'])
                if not attachments:
                    continue
                for att in attachments:
                    filepath = os.path.join(WIKI_UPLOADS_DIR, str(ref['ref_id']), att['filename'])
                    if not os.path.isfile(filepath):
                        continue
                    # Use codename as prefix so _seed_wiki_images() can fuzzy-match
                    codename = ref['codename'].strip().replace('/', '_').replace('\\', '_')
                    ext = att['original_name'].rsplit('.', 1)[-1] if '.' in att['original_name'] else 'png'
                    # If multiple images per product, append a suffix
                    arc_name = f'{codename}.{ext}' if len(attachments) == 1 else f'{codename}_{att["attachment_id"]}.{ext}'
                    inner.write(filepath, arc_name)
                    img_count += 1

        if img_count:
            outer.writestr('printer_images.zip', img_buf.getvalue())

    buf.seek(0)
    app_logger.info('Product reference ZIP export: %d refs, %d images by=%s', len(refs), img_count, current_username())
    return Response(buf.getvalue(), mimetype='application/zip',
                    headers={'Content-Disposition': 'attachment; filename=product_reference_bundle.zip'})


# ---------------------------------------------------------------------------
# Product Wiki — community notes per product
# ---------------------------------------------------------------------------

WIKI_UPLOADS_DIR = os.path.join(DATA_DIR, 'wiki_uploads')
DEVICE_UPLOADS_DIR = os.path.join(DATA_DIR, 'device_uploads')
ALLOWED_EXTENSIONS = {
    'png', 'jpg', 'jpeg', 'gif', 'bmp', 'svg', 'webp',
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'txt',
    'zip', 'tar', 'gz', 'pptx', 'log',
}
MAX_UPLOAD_SIZE = 25 * 1024 * 1024  # 25 MB


# Raster image extensions that should be auto-converted to JPG on upload
_CONVERTIBLE_IMAGE_EXTS = {'png', 'bmp', 'gif', 'webp'}


def _convert_image_to_jpg(data):
    """Convert raster image bytes (PNG, BMP, GIF, WebP) to JPG.
    Returns (jpg_bytes, True) on success, or (original_data, False) on failure."""
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(data))
        if img.mode in ('RGBA', 'LA', 'PA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[-1])
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        buf = io.BytesIO()
        img.save(buf, 'JPEG', quality=85, optimize=True)
        return buf.getvalue(), True
    except Exception:
        return data, False


@app.route('/wiki/<int:ref_id>')
def product_wiki(ref_id):
    """View/edit the wiki page for a product."""
    refs = db.get_all_product_references()
    ref = None
    for r in refs:
        if r['ref_id'] == ref_id:
            ref = r
            break
    if not ref:
        flash('Product not found.', 'error')
        return redirect(url_for('product_reference_list'))
    wiki = db.get_wiki_by_ref_id(ref_id)
    content = wiki['content'] if wiki else ''
    updated_by = wiki['updated_by'] if wiki else ''
    updated_at = wiki['updated_at'] if wiki else ''
    attachments = db.get_wiki_attachments(ref_id)
    predecessor_ref_id = None
    if ref.get('predecessor'):
        for r in refs:
            if r['codename'] == ref['predecessor']:
                predecessor_ref_id = r['ref_id']
                break
    wiki_notes = db.get_wiki_notes(ref_id)
    return render_template('product_wiki.html', ref=ref, content=content,
                           updated_by=updated_by, updated_at=updated_at,
                           attachments=attachments,
                           predecessor_ref_id=predecessor_ref_id,
                           wiki_notes=wiki_notes)


@app.route('/wiki/<int:ref_id>/save', methods=['POST'])
@permission_required('wiki')
def product_wiki_save(ref_id):
    """Save wiki content."""
    content = request.form.get('content', '')
    username = g.user['username'] if g.user else 'guest'
    db.save_wiki(ref_id, content, updated_by=username)
    app_logger.info('Wiki saved: ref_id=%d by=%s', ref_id, current_username())
    flash('Wiki saved.', 'success')
    return redirect(url_for('product_wiki', ref_id=ref_id))


@app.route('/wiki/<int:ref_id>/upload', methods=['POST'])
@permission_required('wiki')
def wiki_upload(ref_id):
    """Upload an attachment to a product wiki."""

    file = request.files.get('attachment')
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('product_wiki', ref_id=ref_id))

    original_name = file.filename
    ext = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else ''
    if ext not in ALLOWED_EXTENSIONS:
        flash(f'File type .{ext} is not allowed.', 'error')
        return redirect(url_for('product_wiki', ref_id=ref_id))

    # Read file and check size
    data = file.read()
    if len(data) > MAX_UPLOAD_SIZE:
        flash('File exceeds 25 MB limit.', 'error')
        return redirect(url_for('product_wiki', ref_id=ref_id))

    # Auto-convert raster images (PNG, BMP, GIF, WebP) to JPG to save space
    if ext in _CONVERTIBLE_IMAGE_EXTS:
        jpg_data, converted = _convert_image_to_jpg(data)
        if converted:
            data = jpg_data
            ext = 'jpg'
            original_name = original_name.rsplit('.', 1)[0] + '.jpg'

    # Save to disk with unique filename
    upload_dir = os.path.join(WIKI_UPLOADS_DIR, str(ref_id))
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f'{uuid.uuid4().hex}.{ext}'
    filepath = os.path.join(upload_dir, safe_name)
    with open(filepath, 'wb') as f:
        f.write(data)

    content_type = 'image/jpeg' if ext in ('jpg', 'jpeg') else (file.content_type or '')
    db.add_wiki_attachment(
        ref_id=ref_id,
        filename=safe_name,
        original_name=original_name,
        content_type=content_type,
        size_bytes=len(data),
        uploaded_by=g.user['username'] if g.user else 'guest',
    )
    app_logger.info('Wiki attachment uploaded: ref_id=%d file="%s" (%d bytes) by=%s',
                    ref_id, original_name, len(data), current_username())
    flash(f'Uploaded {original_name}.', 'success')
    return redirect(url_for('product_wiki', ref_id=ref_id))


@app.route('/wiki/attachment/<int:attachment_id>')
def wiki_download(attachment_id):
    """Download a wiki attachment (public)."""
    att = db.get_wiki_attachment(attachment_id)
    if not att:
        return 'Attachment not found', 404
    filepath = os.path.join(WIKI_UPLOADS_DIR, str(att['ref_id']), att['filename'])
    if not os.path.isfile(filepath):
        return 'File not found on disk', 404
    return send_file(filepath, download_name=att['original_name'], as_attachment=True)


@app.route('/wiki/attachment/<int:attachment_id>/preview')
def wiki_attachment_preview(attachment_id):
    """Serve an attachment inline for image preview (public)."""
    att = db.get_wiki_attachment(attachment_id)
    if not att:
        return 'Attachment not found', 404
    filepath = os.path.join(WIKI_UPLOADS_DIR, str(att['ref_id']), att['filename'])
    if not os.path.isfile(filepath):
        return 'File not found on disk', 404
    return send_file(filepath, mimetype=att['content_type'])


@app.route('/wiki/attachment/<int:attachment_id>/delete', methods=['POST'])
@permission_required('wiki')
def wiki_delete_attachment(attachment_id):
    """Delete a wiki attachment."""
    att = db.get_wiki_attachment(attachment_id)
    if not att:
        flash('Attachment not found.', 'error')
        return redirect(url_for('product_reference_list'))
    # Delete file from disk
    filepath = os.path.join(WIKI_UPLOADS_DIR, str(att['ref_id']), att['filename'])
    if os.path.isfile(filepath):
        os.remove(filepath)
    db.delete_wiki_attachment(attachment_id)
    app_logger.info('Wiki attachment deleted: id=%d file="%s" ref_id=%d by=%s',
                    attachment_id, att['original_name'], att['ref_id'], current_username())
    flash(f'Deleted {att["original_name"]}.', 'success')
    return redirect(url_for('product_wiki', ref_id=att['ref_id']))


@app.route('/wiki/<int:ref_id>/notes', methods=['POST'])
@permission_required('wiki')
def add_wiki_note(ref_id):
    """Add a note to a product wiki."""
    content = request.form.get('note_content', '').strip()
    if not content:
        flash('Note cannot be empty.', 'error')
        return redirect(url_for('product_wiki', ref_id=ref_id))
    if len(content) > 2000:
        flash('Note is too long (max 2000 characters).', 'error')
        return redirect(url_for('product_wiki', ref_id=ref_id))
    author = current_username() if g.user else request.form.get('author_name', '').strip() or 'Anonymous'
    db.add_wiki_note(ref_id, author, content)
    app_logger.info('Wiki note added: ref_id=%d by=%s', ref_id, author)
    flash('Note added.', 'success')
    return redirect(url_for('product_wiki', ref_id=ref_id))


@app.route('/wiki/notes/<int:note_id>/delete', methods=['POST'])
@permission_required('wiki')
def delete_wiki_note_route(note_id):
    """Delete a wiki note. Requires a logged-in user."""
    ref_id = request.form.get('ref_id', type=int)
    if not g.user:
        flash('You must be logged in to delete notes.', 'error')
        return redirect(url_for('product_wiki', ref_id=ref_id))
    db.delete_wiki_note(note_id)
    app_logger.info('Wiki note deleted: note_id=%d by=%s', note_id, current_username())
    flash('Note deleted.', 'success')
    return redirect(url_for('product_wiki', ref_id=ref_id))


@app.route('/wiki/repair', methods=['POST'])
@permission_required('wiki')
def wiki_repair_attachments():
    """Manually run attachment integrity check — removes orphaned DB records."""
    result = db.check_attachment_integrity(WIKI_UPLOADS_DIR)
    if result['orphaned_removed'] > 0:
        app_logger.info('Manual attachment repair: removed %d orphaned records by=%s',
                        result['orphaned_removed'], current_username())
        flash(f'Repair complete: removed {result["orphaned_removed"]} broken attachment references.', 'success')
    else:
        flash(f'All {result["total_checked"]} attachments are intact. No repairs needed.', 'success')
    return redirect(request.referrer or url_for('product_reference_list'))


# ---------------------------------------------------------------------------
# Device Attachments — any logged-in user can upload, devices perm to delete
# ---------------------------------------------------------------------------

@app.route('/devices/<device_id>/upload', methods=['POST'])
def device_upload(device_id):
    """Upload an attachment to a device. Anyone can upload."""
    device = db.get_device(device_id)
    if not device:
        flash('Device not found.', 'error')
        return redirect(url_for('device_list'))

    file = request.files.get('attachment')
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('device_detail', device_id=device_id))

    original_name = file.filename
    ext = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else ''
    if ext not in ALLOWED_EXTENSIONS:
        flash(f'File type .{ext} is not allowed.', 'error')
        return redirect(url_for('device_detail', device_id=device_id))

    data = file.read()
    if len(data) > MAX_UPLOAD_SIZE:
        flash('File exceeds 25 MB limit.', 'error')
        return redirect(url_for('device_detail', device_id=device_id))

    # Auto-convert raster images (PNG, BMP, GIF, WebP) to JPG to save space
    if ext in _CONVERTIBLE_IMAGE_EXTS:
        jpg_data, converted = _convert_image_to_jpg(data)
        if converted:
            data = jpg_data
            ext = 'jpg'
            original_name = original_name.rsplit('.', 1)[0] + '.jpg'

    upload_dir = os.path.join(DEVICE_UPLOADS_DIR, str(device_id))
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f'{uuid.uuid4().hex}.{ext}'
    filepath = os.path.join(upload_dir, safe_name)
    with open(filepath, 'wb') as f:
        f.write(data)

    content_type = 'image/jpeg' if ext in ('jpg', 'jpeg') else (file.content_type or '')
    db.add_device_attachment(
        device_id=device_id,
        filename=safe_name,
        original_name=original_name,
        content_type=content_type,
        size_bytes=len(data),
        uploaded_by=current_username(),
    )
    app_logger.info('Device attachment uploaded: device_id=%s file="%s" (%d bytes) by=%s',
                    device_id, original_name, len(data), current_username())
    flash(f'Uploaded {original_name}.', 'success')
    return redirect(url_for('device_detail', device_id=device_id))


@app.route('/device/attachment/<int:attachment_id>')
def device_download(attachment_id):
    """Download a device attachment (public)."""
    att = db.get_device_attachment(attachment_id)
    if not att:
        return 'Attachment not found', 404
    filepath = os.path.join(DEVICE_UPLOADS_DIR, str(att['device_id']), att['filename'])
    if not os.path.isfile(filepath):
        return 'File not found on disk', 404
    return send_file(filepath, download_name=att['original_name'], as_attachment=True)


@app.route('/device/attachment/<int:attachment_id>/preview')
def device_attachment_preview(attachment_id):
    """Serve a device attachment inline for image preview (public)."""
    att = db.get_device_attachment(attachment_id)
    if not att:
        return 'Attachment not found', 404
    filepath = os.path.join(DEVICE_UPLOADS_DIR, str(att['device_id']), att['filename'])
    if not os.path.isfile(filepath):
        return 'File not found on disk', 404
    return send_file(filepath, mimetype=att['content_type'])


@app.route('/device/attachment/<int:attachment_id>/delete', methods=['POST'])
@permission_required('devices')
def device_delete_attachment(attachment_id):
    """Delete a device attachment (requires devices permission)."""
    att = db.get_device_attachment(attachment_id)
    if not att:
        flash('Attachment not found.', 'error')
        return redirect(url_for('device_list'))
    filepath = os.path.join(DEVICE_UPLOADS_DIR, str(att['device_id']), att['filename'])
    if os.path.isfile(filepath):
        os.remove(filepath)
    db.delete_device_attachment(attachment_id)
    app_logger.info('Device attachment deleted: id=%d file="%s" device_id=%s by=%s',
                    attachment_id, att['original_name'], att['device_id'], current_username())
    flash(f'Deleted {att["original_name"]}.', 'success')
    return redirect(url_for('device_detail', device_id=att['device_id']))


@app.route('/api/devices/distinct/<field>')
def api_distinct_values(field):
    """Return distinct values for a device field, for autocomplete."""
    values = db.get_distinct_values(field)
    return jsonify(values)


@app.route('/api/reference/search')
def api_reference_search():
    """JSON API for printer dropdown in the device form."""
    q = request.args.get('q', '').strip()
    refs = db.get_all_product_references(q)
    return jsonify([{
        'ref_id': r['ref_id'],
        'codename': r['codename'],
        'model_name': r['model_name'],
        'wifi_gen': r['wifi_gen'],
        'year': r['year'],
        'chip_manufacturer': r['chip_manufacturer'],
        'chip_codename': r['chip_codename'],
        'fw_codebase': r['fw_codebase'],
        'print_technology': r['print_technology'],
        'cartridge_toner': r.get('cartridge_toner', ''),
        'predecessor': r.get('predecessor', ''),
    } for r in refs])


# ---------------------------------------------------------------------------
# Health check endpoint (public, no auth required)
# ---------------------------------------------------------------------------

@app.route('/health')
def health_check():
    """Return backup and database health status as JSON for external monitoring."""
    health = db.get_backup_health()
    db_status = db.get_database_status()
    health['database'] = {
        'integrity': db_status['integrity'],
        'size_bytes': db_status['size_bytes'],
        'wal_size_bytes': db_status['wal_size_bytes'],
        'table_counts': db_status['table_counts'],
    }
    if db_status['integrity'] != 'ok':
        health['healthy'] = False
        health['issues'].append(f'Database integrity check failed: {db_status["integrity"]}')
    status_code = 200 if health['healthy'] else 503
    return jsonify(health), status_code


# ---------------------------------------------------------------------------
# Favicon / Apple Touch Icon (generated in-memory to suppress browser 404s)
# ---------------------------------------------------------------------------

_FAVICON_SVG = (
    '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">'
    '<rect width="32" height="32" rx="6" fill="#3b82f6"/>'
    '<text x="16" y="23" text-anchor="middle" fill="white" '
    'font-size="20" font-family="sans-serif" font-weight="bold">I</text></svg>'
)


@app.route('/favicon.ico')
def favicon():
    return Response(_FAVICON_SVG, mimetype='image/svg+xml',
                    headers={'Cache-Control': 'public, max-age=86400'})


@app.route('/apple-touch-icon.png')
@app.route('/apple-touch-icon-precomposed.png')
@app.route('/apple-touch-icon-<dimensions>.png')
@app.route('/apple-touch-icon-<dimensions>-precomposed.png')
def apple_touch_icon(**kwargs):
    return Response(_FAVICON_SVG, mimetype='image/svg+xml',
                    headers={'Cache-Control': 'public, max-age=86400'})


# ---------------------------------------------------------------------------
# Global error handlers
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def not_found(e):
    app_logger.warning('404 Not Found: path=%s method=%s ip=%s user=%s user_agent=%s',
                       request.path, request.method, request.remote_addr,
                       current_username(), request.user_agent.string[:120])
    flash('Page not found.', 'error')
    return redirect(url_for('dashboard'))


@app.errorhandler(500)
def internal_error(e):
    tb = traceback.format_exc()
    app_logger.error('500 Internal Server Error: path=%s method=%s ip=%s user=%s\n'
                     'Exception: %s\nTraceback:\n%s',
                     request.path, request.method, request.remote_addr,
                     current_username(), e, tb)
    flash('An unexpected error occurred.', 'error')
    return redirect(url_for('dashboard'))


@app.errorhandler(Exception)
def unhandled_exception(e):
    tb = traceback.format_exc()
    app_logger.error('Unhandled exception: path=%s method=%s ip=%s user=%s\n'
                     'Exception type: %s — %s\nTraceback:\n%s',
                     request.path, request.method, request.remote_addr,
                     current_username(), type(e).__name__, e, tb)
    flash('An unexpected error occurred.', 'error')
    return redirect(url_for('dashboard'))

# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def _safe_print(*args, **kwargs):
    """print() wrapper that won't crash when stdout is None (windowed PyInstaller)."""
    try:
        if sys.stdout is not None:
            print(*args, **kwargs)
    except Exception:
        pass


if __name__ == '__main__':
    # In windowed PyInstaller builds, sys.stdout/stderr are None; redirect to
    # devnull so any incidental print/traceback doesn't crash the process.
    if sys.stdout is None:
        sys.stdout = open(os.devnull, 'w')
    if sys.stderr is None:
        sys.stderr = open(os.devnull, 'w')

    server_cfg = _load_server_config()
    default_port = server_cfg.get('port', 8080)
    default_host = server_cfg.get('host', '0.0.0.0')

    parser = argparse.ArgumentParser(description='HP Connectivity Team Inventory System')
    parser.add_argument('--host', default=default_host, help=f'Host to bind to (default: {default_host})')
    parser.add_argument('--port', type=int, default=default_port, help=f'Port to listen on (default: {default_port})')
    parser.add_argument('--dev', action='store_true', help='Run in development mode with debug enabled')
    parser.add_argument('--reset-admin', metavar='PASSWORD',
                        help='Reset admin password to PASSWORD and exit. Requires server access. Creates admin if none exists.')
    parser.add_argument('--export-sql', metavar='FILE', help='Export database to SQL dump file and exit')
    parser.add_argument('--emergency-backup', nargs='?', const=True, metavar='PATH',
                        help='Create an emergency database backup and exit')
    parser.add_argument('--convert-png-to-jpg', action='store_true',
                        help='Convert all PNG upload images to JPG to save disk space and exit')
    args = parser.parse_args()

    # --- Recovery CLI commands (run and exit) ---
    if args.reset_admin:
        new_pw = args.reset_admin
        if len(new_pw) < 4:
            _safe_print('  ERROR: Password must be at least 4 characters.')
            exit(1)
        db.init_db()
        username, created = db.reset_admin_password(new_pw)
        if created:
            _safe_print(f'  Admin user created: {username}')
        else:
            _safe_print(f'  Password reset for admin user: {username}')
        _safe_print('  You can now log in with the new credentials.')
        exit(0)

    if args.export_sql:
        db.init_db()
        success = db.export_database_to_sql(args.export_sql)
        if success:
            _safe_print(f'  Database exported to: {args.export_sql}')
        else:
            _safe_print('  Export failed. Check logs for details.')
            exit(1)
        exit(0)

    if args.emergency_backup:
        db.init_db()
        dest = args.emergency_backup if args.emergency_backup is not True else None
        path = db.emergency_backup(dest)
        _safe_print(f'  Emergency backup created: {path}')
        exit(0)

    if args.convert_png_to_jpg:
        db.init_db()
        _safe_print('  Converting PNG uploads to JPG...')
        stats = db.convert_png_uploads_to_jpg()
        _safe_print(f'  Converted: {stats["converted"]}')
        _safe_print(f'  Skipped:   {stats["skipped"]}')
        _safe_print(f'  Errors:    {stats["errors"]}')
        _safe_print(f'  Saved:     {stats["bytes_saved"] // 1024} KB')
        exit(0)

    url = f'http://{args.host}:{args.port}'
    mode = 'DEVELOPMENT' if args.dev else 'PRODUCTION'
    _version = _app_version

    # --- Single-instance check -------------------------------------------
    # If another copy of the app is already serving on this port, open the
    # browser to the running instance and exit immediately so we don't end
    # up with a confusing "address already in use" crash.
    def _probe_existing_instance(host, port):
        import socket as _sock
        probe_host = '127.0.0.1' if host in ('0.0.0.0', '::') else host
        try:
            s = _sock.create_connection((probe_host, port), timeout=0.5)
            s.close()
            return True
        except (OSError, _sock.timeout):
            return False

    if not args.dev and _probe_existing_instance(args.host, args.port):
        app_logger.info('Another instance is already running on %s:%s — opening browser and exiting',
                        args.host, args.port)
        _safe_print(f'  Another instance is already running at {url}')
        _safe_print('  Opening browser to the existing instance...')
        try:
            import webbrowser
            browser_host = '127.0.0.1' if args.host in ('0.0.0.0', '::') else args.host
            webbrowser.open(f'http://{browser_host}:{args.port}')
        except Exception:
            pass
        exit(0)
    # ---------------------------------------------------------------------

    w = 49  # inner width between | chars
    _safe_print()
    _safe_print(f'  +{"-" * w}+')
    _safe_print(f'  |{"HP Connectivity Team Inventory System":^{w}}|')
    _safe_print(f'  |{("v" + _version):^{w}}|')
    _safe_print(f'  |{"":^{w}}|')
    _safe_print(f'  |{"  Running at: " + url:<{w}}|')
    _safe_print(f'  |{"  Mode: " + mode:<{w}}|')
    _safe_print(f'  |{"":^{w}}|')
    _safe_print(f'  |{"  Press Ctrl+C to stop":<{w}}|')
    _safe_print(f'  +{"-" * w}+')
    _safe_print()
    # Also log startup to the app log (visible in the /logs viewer)
    app_logger.info('Application v%s starting on %s (%s mode)', _version, url, mode)

    if args.dev:
        app.run(host=args.host, port=args.port, debug=True)
    else:
        try:
            from waitress import serve
        except ImportError as _imp_err:
            app_logger.critical('waitress could not be imported: %s', _imp_err)
            _safe_print(f"  ERROR: waitress could not be imported ({_imp_err}).")
            _safe_print("  The application requires waitress for multithreaded serving.")
            exit(1)
        app_logger.info('Starting production server (waitress, 16 threads) on %s:%s',
                        args.host, args.port)
        serve(app, host=args.host, port=args.port, threads=16,
              ident='HP Connectivity Inventory')
